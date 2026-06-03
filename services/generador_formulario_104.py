"""
Generador Formulario 104 - Declaración del Impuesto al Valor Agregado (IVA)
Formato: SRI Ecuador
Período: Mensual
"""
from models import db
from models.user import Factura, SaldoIVAMes
from services.credito_tributario import CreditoTributario
from datetime import datetime, timedelta
from sqlalchemy import extract
import json


class GeneradorFormulario104:
    """Generador del Formulario 104 SRI - Detalles IVA"""

    @staticmethod
    def obtener_datos_declaracion(usuario_id, anio, mes):
        """Obtiene todos los datos para generar el Formulario 104

        Args:
            usuario_id: ID del usuario
            anio: Año fiscal
            mes: Mes fiscal (1-12)

        Returns:
            Dict con datos del formulario
        """
        # Obtener saldo IVA del mes
        saldo = SaldoIVAMes.query.filter_by(
            usuario_id=usuario_id,
            anio=anio,
            mes=mes
        ).first()

        if not saldo:
            # Calcular si no existe
            saldo_dict = CreditoTributario.calcular_saldo_iva_mes(
                usuario_id, anio, mes
            )
        else:
            saldo_dict = {
                'anio': saldo.anio,
                'mes': saldo.mes,
                'iva_cobrado': float(saldo.iva_cobrado or 0),
                'iva_pagado': float(saldo.iva_pagado or 0),
                'saldo_anterior': float(saldo.saldo_anterior or 0),
                'saldo_final': float(saldo.saldo_final or 0),
            }

        # Obtener detalles por tarifa de VENTAS (ingresos)
        ventas = db.session.query(
            db.func.sum(Factura.base_iva).label('base_total'),
            db.func.sum(Factura.valor_iva).label('iva_total'),
        ).filter(
            Factura.usuario_id == usuario_id,
            Factura.tipo == 'ingreso',
            extract('year', Factura.fecha_emision) == anio,
            extract('month', Factura.fecha_emision) == mes
        ).first()

        # Obtener detalles por tarifa de COMPRAS (gastos)
        compras = db.session.query(
            db.func.sum(Factura.base_iva).label('base_total'),
            db.func.sum(Factura.valor_iva).label('iva_total'),
        ).filter(
            Factura.usuario_id == usuario_id,
            Factura.tipo == 'gasto',
            extract('year', Factura.fecha_emision) == anio,
            extract('month', Factura.fecha_emision) == mes
        ).first()

        return {
            'periodo': {
                'anio': anio,
                'mes': mes,
                'fecha_presentacion': datetime.now().isoformat(),
            },
            'ventas': {
                'base_iva': float(ventas.base_total or 0) if ventas else 0,
                'iva_cobrado': float(ventas.iva_total or 0) if ventas else 0,
            },
            'compras': {
                'base_iva': float(compras.base_total or 0) if compras else 0,
                'iva_pagado': float(compras.iva_total or 0) if compras else 0,
            },
            'credito': saldo_dict,
            'resumen': {
                'iva_cobrado_periodo': float(ventas.iva_total or 0) if ventas else 0,
                'iva_pagado_periodo': float(compras.iva_total or 0) if compras else 0,
                'saldo_iva_mes': saldo_dict['saldo_final'],
                'debe_pagar': saldo_dict['saldo_final'] < 0,
                'tiene_credito': saldo_dict['saldo_final'] > 0,
            }
        }

    @staticmethod
    def generar_excel(usuario_id, anio, mes):
        """Genera Formulario 104 en formato Excel

        Returns:
            BytesIO object con archivo Excel
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise ImportError("openpyxl no instalado. Instalar: pip install openpyxl")

        datos = GeneradorFormulario104.obtener_datos_declaracion(usuario_id, anio, mes)

        wb = Workbook()
        ws = wb.active
        ws.title = f"Formulario 104 - {anio}/{mes:02d}"

        # Estilos
        header_fill = PatternFill(start_color="0D1B2E", end_color="0D1B2E", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Título
        ws['A1'] = "FORMULARIO 104 - DECLARACIÓN DEL IMPUESTO AL VALOR AGREGADO (IVA)"
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')

        # Período
        ws['A3'] = "PERÍODO FISCAL:"
        ws['B3'] = f"{anio}/{mes:02d}"
        ws['A3'].font = Font(bold=True)

        # SECCIÓN 1: VENTAS (Ingresos)
        ws['A6'] = "SECCIÓN 1: VENTAS REALIZADAS"
        ws['A6'].font = header_font
        ws['A6'].fill = header_fill

        headers = ['Concepto', 'Base IVA', 'IVA Cobrado']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        row = 8
        ws[f'A{row}'] = "Ventas a tarifa 0%"
        ws[f'A{row}'].border = border
        row += 1
        ws[f'A{row}'] = "Ventas a tarifa 5%"
        ws[f'A{row}'].border = border
        row += 1
        ws[f'A{row}'] = "Ventas a tarifa 12%"
        ws[f'A{row}'].border = border
        row += 1
        ws[f'A{row}'] = "Ventas a tarifa 15%"
        ws[f'A{row}'].border = border
        row += 1

        ws[f'A{row}'] = "TOTAL VENTAS"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = datos['ventas']['base_iva']
        ws[f'C{row}'] = datos['ventas']['iva_cobrado']
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'C{row}'].number_format = '#,##0.00'

        # SECCIÓN 2: COMPRAS (Gastos)
        ws['A13'] = "SECCIÓN 2: COMPRAS REALIZADAS"
        ws['A13'].font = header_font
        ws['A13'].fill = header_fill

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=14, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        row = 15
        ws[f'A{row}'] = "Compras a tarifa 0%"
        ws[f'A{row}'].border = border
        row += 1
        ws[f'A{row}'] = "Compras a tarifa 5%"
        ws[f'A{row}'].border = border
        row += 1
        ws[f'A{row}'] = "Compras a tarifa 12%"
        ws[f'A{row}'].border = border
        row += 1
        ws[f'A{row}'] = "Compras a tarifa 15%"
        ws[f'A{row}'].border = border
        row += 1

        ws[f'A{row}'] = "TOTAL COMPRAS"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = datos['compras']['base_iva']
        ws[f'C{row}'] = datos['compras']['iva_pagado']
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'C{row}'].number_format = '#,##0.00'

        # SECCIÓN 3: CRÉDITO TRIBUTARIO
        ws['A22'] = "SECCIÓN 3: CRÉDITO TRIBUTARIO"
        ws['A22'].font = header_font
        ws['A22'].fill = header_fill

        row = 23
        concepts = [
            ('IVA Cobrado (Ventas)', datos['credito']['iva_cobrado']),
            ('Menos: IVA Pagado (Compras)', -datos['credito']['iva_pagado']),
            ('Saldo Anterior', datos['credito']['saldo_anterior']),
            ('SALDO FINAL DEL MES', datos['credito']['saldo_final']),
        ]

        for concepto, valor in concepts:
            ws[f'A{row}'] = concepto
            ws[f'B{row}'] = valor
            if concepto == 'SALDO FINAL DEL MES':
                ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
                ws[f'A{row}'].fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
                ws[f'B{row}'].font = Font(bold=True, size=11)
                ws[f'B{row}'].fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
            ws[f'B{row}'].number_format = '#,##0.00'
            row += 1

        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

        # Guardar a BytesIO
        import io
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def generar_json(usuario_id, anio, mes):
        """Genera Formulario 104 en formato JSON

        Returns:
            Dict con datos del formulario
        """
        datos = GeneradorFormulario104.obtener_datos_declaracion(usuario_id, anio, mes)
        return {
            'tipo': 'Formulario 104',
            'version': '1.0',
            'periodo': datos['periodo'],
            'secciones': {
                'ventas': datos['ventas'],
                'compras': datos['compras'],
                'credito_tributario': datos['credito'],
            },
            'resumen': datos['resumen'],
        }

    @staticmethod
    def generar_xml(usuario_id, anio, mes):
        """Genera Formulario 104 en formato XML (para SRI)

        Returns:
            String con XML del formulario
        """
        datos = GeneradorFormulario104.obtener_datos_declaracion(usuario_id, anio, mes)

        xml = f"""<?xml version="1.0" encoding="UTF-8"?>
<Formulario104>
    <Periodo>
        <Anio>{datos['periodo']['anio']}</Anio>
        <Mes>{datos['periodo']['mes']:02d}</Mes>
        <FechaPresentacion>{datos['periodo']['fecha_presentacion']}</FechaPresentacion>
    </Periodo>
    <Ventas>
        <BaseIVA>{datos['ventas']['base_iva']:.2f}</BaseIVA>
        <IVACobrado>{datos['ventas']['iva_cobrado']:.2f}</IVACobrado>
    </Ventas>
    <Compras>
        <BaseIVA>{datos['compras']['base_iva']:.2f}</BaseIVA>
        <IVAPagado>{datos['compras']['iva_pagado']:.2f}</IVAPagado>
    </Compras>
    <CreditoTributario>
        <IVACobrado>{datos['credito']['iva_cobrado']:.2f}</IVACobrado>
        <IVAPagado>{datos['credito']['iva_pagado']:.2f}</IVAPagado>
        <SaldoAnterior>{datos['credito']['saldo_anterior']:.2f}</SaldoAnterior>
        <SaldoFinal>{datos['credito']['saldo_final']:.2f}</SaldoFinal>
        <Estado>{"Crédito" if datos['credito']['saldo_final'] > 0 else "Deuda" if datos['credito']['saldo_final'] < 0 else "Neto"}</Estado>
    </CreditoTributario>
    <Resumen>
        <DebePagar>{str(datos['resumen']['debe_pagar']).lower()}</DebePagar>
        <TieneCredito>{str(datos['resumen']['tiene_credito']).lower()}</TieneCredito>
    </Resumen>
</Formulario104>
"""
        return xml
