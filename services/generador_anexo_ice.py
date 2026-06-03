"""
Generador Anexo ICE/PVP - Impuesto a Consumos Especiales
Formato: SRI Ecuador
Período: Mensual
"""
from models import db
from models.user import Factura
from datetime import datetime
from sqlalchemy import extract
import json


class GeneradorAnexoICE:
    """Generador del Anexo ICE/PVP SRI"""

    # Categorías de productos ICE según SRI
    CATEGORIAS_ICE = {
        'bebidas_alcoholicas': {'nombre': 'Bebidas Alcohólicas', 'codigo': '01'},
        'bebidas_no_alcoholicas': {'nombre': 'Bebidas no Alcohólicas', 'codigo': '02'},
        'tabacos': {'nombre': 'Tabacos', 'codigo': '03'},
        'combustibles': {'nombre': 'Combustibles', 'codigo': '04'},
        'vehiculos': {'nombre': 'Vehículos', 'codigo': '05'},
        'cosmeticos': {'nombre': 'Cosméticos', 'codigo': '06'},
        'servicios_telecomunicacion': {'nombre': 'Telecom', 'codigo': '07'},
        'otros': {'nombre': 'Otros', 'codigo': '99'},
    }

    @staticmethod
    def obtener_datos_anexo(usuario_id, anio, mes):
        """Obtiene datos del Anexo ICE/PVP

        Args:
            usuario_id: ID del usuario
            anio: Año fiscal
            mes: Mes fiscal (1-12)

        Returns:
            Dict con datos del anexo ICE
        """
        # Obtener facturas con ICE del período
        facturas_ice = db.session.query(
            Factura.id,
            Factura.numero_factura,
            Factura.ruc_proveedor,
            Factura.descripcion,
            Factura.base_iva,
            Factura.valor_iva,
            Factura.notas_auditoria,
            Factura.fecha_emision,
            Factura.tipo,
        ).filter(
            Factura.usuario_id == usuario_id,
            extract('year', Factura.fecha_emision) == anio,
            extract('month', Factura.fecha_emision) == mes,
            # Filtrar facturas que tengan ICE (buscar en notas_auditoria)
        ).all()

        # Procesar y agrupar por categoría
        resumen_ice = {
            'periodo': {'anio': anio, 'mes': mes},
            'fecha_generacion': datetime.now().isoformat(),
            'total_registros': 0,
            'total_base_imponible': 0.0,
            'total_ice': 0.0,
            'categorias': {},
            'detalles': []
        }

        for factura in facturas_ice:
            try:
                # Intentar extraer datos de notas_auditoria
                datos_auditoria = json.loads(factura.notas_auditoria or '{}')
                categoria = datos_auditoria.get('categoria_ice', 'otros')
                tasa_ice = float(datos_auditoria.get('tasa_ice', 0))
                valor_ice = float(datos_auditoria.get('valor_ice', 0))
            except:
                categoria = 'otros'
                tasa_ice = 0
                valor_ice = 0

            # Agrupar por categoría
            if categoria not in resumen_ice['categorias']:
                resumen_ice['categorias'][categoria] = {
                    'codigo': GeneradorAnexoICE.CATEGORIAS_ICE.get(
                        categoria, {'codigo': '99'}
                    )['codigo'],
                    'nombre': GeneradorAnexoICE.CATEGORIAS_ICE.get(
                        categoria, {'nombre': 'Otros'}
                    )['nombre'],
                    'cantidad': 0,
                    'base_imponible': 0.0,
                    'tasa_promedio': 0.0,
                    'valor_ice': 0.0,
                }

            resumen_ice['categorias'][categoria]['cantidad'] += 1
            resumen_ice['categorias'][categoria]['base_imponible'] += float(
                factura.base_iva or 0
            )
            resumen_ice['categorias'][categoria]['valor_ice'] += valor_ice

            # Agregar detalle
            resumen_ice['detalles'].append({
                'factura_id': factura.numero_factura,
                'ruc_proveedor': factura.ruc_proveedor,
                'descripcion': factura.descripcion[:50],
                'base_imponible': float(factura.base_iva or 0),
                'tasa_ice': tasa_ice,
                'valor_ice': valor_ice,
                'tipo': factura.tipo,
                'fecha': factura.fecha_emision.isoformat() if factura.fecha_emision else '',
            })

            resumen_ice['total_registros'] += 1
            resumen_ice['total_base_imponible'] += float(factura.base_iva or 0)
            resumen_ice['total_ice'] += valor_ice

        return resumen_ice

    @staticmethod
    def generar_excel(usuario_id, anio, mes):
        """Genera Anexo ICE en formato Excel

        Returns:
            BytesIO object con archivo Excel
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise ImportError("openpyxl no instalado. Instalar: pip install openpyxl")

        datos = GeneradorAnexoICE.obtener_datos_anexo(usuario_id, anio, mes)

        wb = Workbook()
        ws = wb.active
        ws.title = f"Anexo ICE - {anio}/{mes:02d}"

        # Estilos
        header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Título
        ws['A1'] = "ANEXO ICE/PVP - IMPUESTO A CONSUMOS ESPECIALES"
        ws['A1'].font = title_font
        ws.merge_cells('A1:E1')

        # Período
        ws['A3'] = "PERÍODO FISCAL:"
        ws['B3'] = f"{datos['periodo']['anio']}/{datos['periodo']['mes']:02d}"
        ws['A3'].font = Font(bold=True)

        # SECCIÓN 1: RESUMEN POR CATEGORÍA
        ws['A6'] = "SECCIÓN 1: RESUMEN POR CATEGORÍA"
        ws['A6'].font = header_font
        ws['A6'].fill = header_fill

        headers = ['Categoría', 'Código', 'Cantidad', 'Base Imponible', 'Tasa', 'Valor ICE']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        row = 8
        for categoria, datos_cat in datos['categorias'].items():
            ws[f'A{row}'] = datos_cat['nombre']
            ws[f'B{row}'] = datos_cat['codigo']
            ws[f'C{row}'] = datos_cat['cantidad']
            ws[f'D{row}'] = datos_cat['base_imponible']
            ws[f'E{row}'] = '0%'
            ws[f'F{row}'] = datos_cat['valor_ice']

            for col in range(1, 7):
                ws.cell(row=row, column=col).border = border
                if col in [4, 6]:
                    ws.cell(row=row, column=col).number_format = '#,##0.00'

            row += 1

        # Totales
        ws[f'A{row}'] = "TOTAL"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'D{row}'] = datos['total_base_imponible']
        ws[f'F{row}'] = datos['total_ice']
        ws[f'D{row}'].font = Font(bold=True)
        ws[f'F{row}'].font = Font(bold=True)
        ws[f'D{row}'].number_format = '#,##0.00'
        ws[f'F{row}'].number_format = '#,##0.00'

        # SECCIÓN 2: DETALLE DE TRANSACCIONES
        ws[f'A{row + 3}'] = "SECCIÓN 2: DETALLE DE TRANSACCIONES"
        ws[f'A{row + 3}'].font = header_font
        ws[f'A{row + 3}'].fill = header_fill

        detail_headers = ['Factura', 'RUC', 'Descripción', 'Base', 'ICE', 'Tipo']
        for col, header in enumerate(detail_headers, 1):
            cell = ws.cell(row=row + 4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        detail_row = row + 5
        for detalle in datos['detalles'][:50]:  # Max 50 detalles
            ws[f'A{detail_row}'] = detalle['factura_id']
            ws[f'B{detail_row}'] = detalle['ruc_proveedor']
            ws[f'C{detail_row}'] = detalle['descripcion']
            ws[f'D{detail_row}'] = detalle['base_imponible']
            ws[f'E{detail_row}'] = detalle['valor_ice']
            ws[f'F{detail_row}'] = detalle['tipo']

            for col in range(1, 7):
                ws.cell(row=detail_row, column=col).border = border
                if col in [4, 5]:
                    ws.cell(row=detail_row, column=col).number_format = '#,##0.00'

            detail_row += 1

        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12

        # Guardar a BytesIO
        import io
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def generar_json(usuario_id, anio, mes):
        """Genera Anexo ICE en formato JSON

        Returns:
            Dict con datos del anexo
        """
        datos = GeneradorAnexoICE.obtener_datos_anexo(usuario_id, anio, mes)
        return {
            'tipo': 'Anexo ICE/PVP',
            'version': '1.0',
            'periodo': datos['periodo'],
            'resumen': {
                'total_registros': datos['total_registros'],
                'total_base_imponible': datos['total_base_imponible'],
                'total_ice': datos['total_ice'],
                'fecha_generacion': datos['fecha_generacion'],
            },
            'categorias': datos['categorias'],
            'detalles': datos['detalles'][:100],  # Max 100 para JSON
        }

    @staticmethod
    def generar_xml(usuario_id, anio, mes):
        """Genera Anexo ICE en formato XML (para SRI)

        Returns:
            String con XML del anexo
        """
        datos = GeneradorAnexoICE.obtener_datos_anexo(usuario_id, anio, mes)

        xml = f"""<?xml version="1.0" encoding="UTF-8"?>
<AnexoICE>
    <Periodo>
        <Anio>{datos['periodo']['anio']}</Anio>
        <Mes>{datos['periodo']['mes']:02d}</Mes>
        <FechaGeneracion>{datos['fecha_generacion']}</FechaGeneracion>
    </Periodo>
    <Resumen>
        <TotalRegistros>{datos['total_registros']}</TotalRegistros>
        <TotalBaseImponible>{datos['total_base_imponible']:.2f}</TotalBaseImponible>
        <TotalICE>{datos['total_ice']:.2f}</TotalICE>
    </Resumen>
    <Categorias>
"""
        for categoria, datos_cat in datos['categorias'].items():
            xml += f"""        <Categoria>
            <Nombre>{datos_cat['nombre']}</Nombre>
            <Codigo>{datos_cat['codigo']}</Codigo>
            <Cantidad>{datos_cat['cantidad']}</Cantidad>
            <BaseImponible>{datos_cat['base_imponible']:.2f}</BaseImponible>
            <ValorICE>{datos_cat['valor_ice']:.2f}</ValorICE>
        </Categoria>
"""
        xml += """    </Categorias>
</AnexoICE>
"""
        return xml
