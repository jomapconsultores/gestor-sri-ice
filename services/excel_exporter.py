import io
from datetime import datetime
from services.ice_calculator import IceCalculator, TAX_DB

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_DISPONIBLE = True
except ImportError:
    EXCEL_DISPONIBLE = False


class ExcelExporter:
    
    @staticmethod
    def crear_estilos():
        return {
            'h_fill': PatternFill(start_color='1a5276', end_color='1a5276', fill_type='solid'),
            'h_font': Font(name='Arial', size=9, bold=True, color='FFFFFF'),
            'total_fill': PatternFill(start_color='27ae60', end_color='27ae60', fill_type='solid'),
            'total_font': Font(name='Arial', size=10, bold=True, color='FFFFFF'),
            'nfont': Font(name='Arial', size=9),
            'bfont': Font(name='Arial', size=9, bold=True),
            'title_font': Font(name='Arial', size=14, bold=True, color='1a5276'),
            'border': Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        }
    
    @staticmethod
    def exportar_declaracion(facturas):
        """Exporta Excel con el detalle de facturas (como SRI-XML.py)"""
        if not EXCEL_DISPONIBLE:
            return None
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Declaracion"
        s = ExcelExporter.crear_estilos()

        encabezados = ['Fecha', 'N Factura', 'RUC Emisor', 'Cliente', 'Subtotal',
                       'Base ICE', 'ICE', 'Base IVA', 'IVA', 'Importe Total']
        
        for j, enc in enumerate(encabezados, 1):
            cell = ws.cell(row=1, column=j, value=enc)
            cell.fill = s['h_fill']
            cell.font = s['h_font']
            cell.border = s['border']
            cell.alignment = Alignment(horizontal='center')
        
        fila = 2
        for f in facturas:
            total = float(f.importe_total or 0)
            ice = float(f.valor_ice or 0)
            iva = float(f.valor_iva or 0)
            vals = [
                f.fecha_emision.strftime('%d/%m/%Y') if f.fecha_emision else '',
                f.numero_factura or '',
                f.ruc_emisor or '',
                f.razon_social_comprador[:40] if f.razon_social_comprador else '',
                round(total - ice - iva, 2),
                float(f.base_ice or 0),
                ice,
                float(f.base_iva or 0),
                iva,
                total,
            ]
            for j, val in enumerate(vals, 1):
                cell = ws.cell(row=fila, column=j, value=val)
                cell.font = s['nfont']
                cell.border = s['border']
                if j >= 5:
                    cell.number_format = '#,##0.00'
            fila += 1
        
        # Totales
        ult = fila
        for j in range(1, 11):
            cell = ws.cell(row=ult, column=j)
            cell.fill = s['total_fill']
            cell.font = s['total_font']
            cell.border = s['border']
        
        ws.cell(row=ult, column=1, value="TOTALES")
        totales = [0.0] * 11
        for f in facturas:
            total = float(f.importe_total or 0)
            ice = float(f.valor_ice or 0)
            iva = float(f.valor_iva or 0)
            totales[5] += round(total - ice - iva, 2)
            totales[6] += float(f.base_ice or 0)
            totales[7] += ice
            totales[8] += float(f.base_iva or 0)
            totales[9] += iva
            totales[10] += total
        for col in [5, 6, 7, 8, 9, 10]:
            ws.cell(row=ult, column=col, value=round(totales[col], 2))
            ws.cell(row=ult, column=col).number_format = '#,##0.00'
        
        for j, w in enumerate([12, 15, 15, 30, 12, 12, 12, 12, 12, 12], 1):
            ws.column_dimensions[get_column_letter(j)].width = w
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def exportar_auditoria(datos_producto, anios_seleccionados):
        """Exporta Excel de auditoria ICE multi-ano (como ICEcapsa.py)"""
        if not EXCEL_DISPONIBLE:
            return None
        
        wb = Workbook()
        s = ExcelExporter.crear_estilos()

        ws = wb.active
        ws.title = "Auditoria ICE"

        titulo = ws.cell(row=1, column=1, value="AUDITORIA ICE - CALCULO POR PRODUCTO")
        titulo.font = s['title_font']
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

        enc = ['Ano', 'Tarifa Esp.', 'Umbral AdV', 'IVA', 'ICE Esp. Unit.',
               'ICE AdV. Unit.', 'ICE Total Unit.', 'Base IVA', 'IVA Total', 'PVP Final']

        for j, e in enumerate(enc, 1):
            cell = ws.cell(row=3, column=j, value=e)
            cell.fill = s['h_fill']
            cell.font = s['h_font']
            cell.border = s['border']

        fila = 4
        for anio in anios_seleccionados:
            if anio == '2024':
                filas_anio = [
                    ('2024 (12%)', IceCalculator.calcular_liquidacion_completa(datos_producto, '2024', iva_tasa=0.12)),
                    ('2024 (15%)', IceCalculator.calcular_liquidacion_completa(datos_producto, '2024', iva_tasa=0.15)),
                ]
            else:
                res = IceCalculator.calcular_liquidacion_completa(datos_producto, anio)
                filas_anio = [(anio, res)]

            for etiqueta, res in filas_anio:
                vals = [
                    etiqueta, res['tarifa_especifica'], res['umbral_advalorem'],
                    f"{int(res['iva_tasa'] * 100)}%", res['ice_especifico_unitario'],
                    res['ice_advalorem_unitario'], res['ice_total_unitario'],
                    res['base_iva'], res['iva_total'], res['pvp']
                ]
                for j, val in enumerate(vals, 1):
                    cell = ws.cell(row=fila, column=j, value=val)
                    cell.font = s['nfont']
                    cell.border = s['border']
                    if j >= 4:
                        cell.number_format = '#,##0.0000' if j <= 7 else '#,##0.00'
                fila += 1

        # Fila de totales (columnas numéricas: ICE esp, ICE adv, ICE total, base IVA, IVA, PVP)
        ult = fila
        for j in range(1, 11):
            cell = ws.cell(row=ult, column=j)
            cell.fill = s['total_fill']
            cell.font = s['total_font']
            cell.border = s['border']
        ws.cell(row=ult, column=1, value="TOTALES")
        _sum = {8: 0.0, 9: 0.0, 10: 0.0}
        for r in range(4, ult):
            for col in (8, 9, 10):
                v = ws.cell(row=r, column=col).value
                if isinstance(v, (int, float)):
                    _sum[col] += v
        for col in [8, 9, 10]:
            ws.cell(row=ult, column=col, value=round(_sum[col], 2))
            ws.cell(row=ult, column=col).number_format = '#,##0.00'

        for j in range(1, 11):
            ws.column_dimensions[get_column_letter(j)].width = 18

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output