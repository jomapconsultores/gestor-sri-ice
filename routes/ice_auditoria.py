"""Sistema Integrado ICE – Auditoría + Anexo SRI (sincronizado con ICEcompleto.py)"""
import io, os, json
from collections import defaultdict
from flask import Blueprint, render_template, request, Response, redirect, url_for, flash
from flask_login import login_required, current_user
from routes.payments import usuario_tiene_modulo
import xml.etree.ElementTree as ET
import xml.dom.minidom as minidom

ice_auditoria = Blueprint('ice_auditoria', __name__)

# ── Base de datos tributaria ──────────────────────────────────────────────────
TAX_DB = {
    "2021": {"esp": 7.18, "art": 1.49, "ind": 13.08, "umb": 4.29, "iva": 0.12},
    "2022": {"esp": 10.00, "art": 1.50, "ind": 13.08, "umb": 4.37, "iva": 0.12},
    "2023": {"esp": 10.00, "art": 1.50, "ind": 13.08, "umb": 4.53, "iva": 0.12},
    "2024": {"esp": 10.15, "art": 1.52, "ind": 13.28, "umb": 4.60, "iva": "MIXTO"},
    "2025": {"esp": 10.30, "art": 1.54, "ind": 13.48, "umb": 4.67, "iva": 0.15},
    "2026": {"esp": 10.41, "art": 1.56, "ind": 13.62, "umb": 4.72, "iva": 0.15},
}

# ── Catálogo con códigos SRI reales ──────────────────────────────────────────
CATALOGO = {
    'LICOR ORO': {
        'codMarca': '019167', 'codProdSRI': '19167', 'presentacion': '13',
        'capacidad': '750', 'unidad': '66', 'grado': '15',
        'codImpuesto': '3031', 'tipo': 'Licor', 'botellas_por_caja': 12
    },
    'LICOR SECO BLANCO': {
        'codMarca': '039919', 'codProdSRI': '39919', 'presentacion': '13',
        'capacidad': '750', 'unidad': '66', 'grado': '15',
        'codImpuesto': '3031', 'tipo': 'Licor', 'botellas_por_caja': 12
    },
    'AGUARDIENTE DE CAÑA': {
        'codMarca': '036886', 'codProdSRI': '36886', 'presentacion': '13',
        'capacidad': '750', 'unidad': '66', 'grado': '15',
        'codImpuesto': '3031', 'tipo': 'Licor', 'botellas_por_caja': 12
    },
    'VODKA SECO GLACIAL': {
        'codMarca': '027298', 'codProdSRI': '27298', 'presentacion': '13',
        'capacidad': '750', 'unidad': '66', 'grado': '15',
        'codImpuesto': '3031', 'tipo': 'Licor', 'botellas_por_caja': 12
    },
    'COCKTAIL CON VODKA SABOR A MARACUYA': {
        'codMarca': '022744', 'codProdSRI': '22744', 'presentacion': '13',
        'capacidad': '800', 'unidad': '66', 'grado': '5',
        'codImpuesto': '3031', 'tipo': 'Cocktail', 'botellas_por_caja': 12
    },
    'COCKTAIL CON BAJO GRADO ALCOHOLICO SABOR A DURAZNO': {
        'codMarca': '006868', 'codProdSRI': '6868', 'presentacion': '13',
        'capacidad': '800', 'unidad': '66', 'grado': '5',
        'codImpuesto': '3031', 'tipo': 'Cocktail', 'botellas_por_caja': 12
    },
    'COCKTAIL CON VODKA SABOR A GUARANA': {
        'codMarca': '039912', 'codProdSRI': '39912', 'presentacion': '13',
        'capacidad': '750', 'unidad': '66', 'grado': '5',
        'codImpuesto': '3031', 'tipo': 'Cocktail', 'botellas_por_caja': 12
    },
}

PALABRAS_CLAVE = {
    'LICOR ORO': ['LICOR ORO'],
    'LICOR SECO BLANCO': ['LICOR SECO BLANCO'],
    'AGUARDIENTE DE CAÑA': ['AGUARDIENTE DE CAÑA', 'AGUARDIENTE'],
    'VODKA SECO GLACIAL': ['VODKA SECO GLACIAL', 'VODKA SECO'],
    'COCKTAIL CON VODKA SABOR A MARACUYA': ['MARACUYA', 'MARACUYÁ'],
    'COCKTAIL CON BAJO GRADO ALCOHOLICO SABOR A DURAZNO': ['DURAZNO'],
    'COCKTAIL CON VODKA SABOR A GUARANA': ['GUARANA', 'GUARANÁ'],
}

CATALOGO_DEFAULT = {
    'codMarca': '000000', 'codProdSRI': '', 'presentacion': '13',
    'capacidad': '750', 'unidad': '66', 'grado': '15',
    'codImpuesto': '3031', 'tipo': 'Licor', 'botellas_por_caja': 12
}

# ── Mapeo tipoIdentificacion factura → tipoIdCliente ICE (XSD acepta letras) ─
_TIPO_ID_ICE = {
    '04': 'R',   # RUC
    '05': 'C',   # Cédula de Identidad
    '06': 'P',   # Pasaporte
    '07': 'F',   # Consumidor Final
    '08': 'F',   # Identificación exterior → Consumidor Final
}


def _mapear_tipo_id(tipo_id_factura):
    return _TIPO_ID_ICE.get(str(tipo_id_factura).strip(), 'F')


def buscar_catalogo(descripcion):
    desc_u = descripcion.upper()
    for nombre, claves in PALABRAS_CLAVE.items():
        if any(c in desc_u for c in claves) and nombre in CATALOGO:
            return CATALOGO[nombre].copy()
    return CATALOGO_DEFAULT.copy()


def _construir_cod_prod_ice(info_cat):
    """Construye el código completo codProdICE para el anexo XML del SRI."""
    cod_sri = info_cat.get('codProdSRI', '').strip()
    if not cod_sri:
        return info_cat.get('codImpuesto', '3031')
    if '-' in cod_sri:
        return cod_sri
    pres = info_cat.get('presentacion', '13').zfill(3)
    cap = info_cat.get('capacidad', '750').zfill(6)
    und = info_cat.get('unidad', '66')
    grad = info_cat.get('grado', '15').zfill(6)
    cimp = info_cat.get('codImpuesto', '3031')
    return f"{cimp}-057-{cod_sri.zfill(6)}-{pres}-{cap}-{und}-593-{grad}"


def es_pack(desc):
    d = desc.upper()
    if 'PACK' in d:
        return True
    if '+' in desc and any(p in d for p in ['AGUARDIENTE', 'VODKA', 'LICOR']):
        return True
    return False


def descomponer_pack(desc):
    d = desc.upper()
    productos = []
    if 'VODKA SECO GLACIAL' in d or 'VODKA SECO' in d or 'VODKA' in d:
        productos.append(('VODKA SECO GLACIAL', '750'))
    if 'LICOR ORO' in d:
        productos.append(('LICOR ORO', '750'))
    if 'AGUARDIENTE' in d:
        cap = '375' if '375' in desc else '750'
        productos.append(('AGUARDIENTE DE CAÑA', cap))
    return productos or [('LICOR ORO', '750')]


def ice_especifico(t_esp, grado, vol_cc):
    return t_esp * (grado / 100.0) * (vol_cc / 1000.0)


def ice_advalorem(precio_bot, vol_cc, umbral):
    p_litro = (precio_bot * 1000.0) / vol_cc if vol_cc > 0 else 0
    if p_litro > umbral:
        return (p_litro - umbral) * 0.75 * (vol_cc / 1000.0)
    return 0.0


def _requiere_modulo():
    if current_user.is_admin:
        return None
    if not usuario_tiene_modulo('ice_auditoria'):
        flash('Requieres el módulo Auditoría ICE Completa para usar esta herramienta.', 'warning')
        return redirect(url_for('payments.ver_planes'))
    return None


# ── Parseo XML de facturas ICE ───────────────────────────────────────────────

def parsear_xml_ice(contenido_bytes):
    """Parsea un XML de factura desde bytes (en memoria, sin guardar a disco)."""
    try:
        root = ET.fromstring(contenido_bytes)
        if root.tag != 'factura':
            return [], []

        fact = root.find('infoFactura')
        if fact is None:
            return [], []

        fecha = (fact.find('fechaEmision').text or '') if fact.find('fechaEmision') is not None else ''
        tipo_id = (fact.find('tipoIdentificacionComprador').text or '') if fact.find('tipoIdentificacionComprador') is not None else ''
        id_cli = (fact.find('identificacionComprador').text or '') if fact.find('identificacionComprador') is not None else ''
        razon_el = fact.find('razonSocialComprador')
        razon = razon_el.text if razon_el is not None else 'CONSUMIDOR FINAL'
        importe_total = float((fact.find('importeTotal').text or '0') if fact.find('importeTotal') is not None else '0')

        registros_ice = []
        registros_pvp = []
        detalles = root.find('detalles')
        if detalles is None:
            return [], []

        for det in detalles.findall('detalle'):
            cod = (det.find('codigoPrincipal').text or '') if det.find('codigoPrincipal') is not None else ''
            desc = (det.find('descripcion').text or '') if det.find('descripcion') is not None else ''
            try:
                cant = float(det.find('cantidad').text or 0)
            except (ValueError, AttributeError):
                cant = 0.0
            try:
                p_unit = float(det.find('precioUnitario').text or 0)
            except (ValueError, AttributeError):
                p_unit = 0.0
            try:
                p_total = float(det.find('precioTotalSinImpuesto').text or 0)
            except (ValueError, AttributeError):
                p_total = 0.0

            ice_val = iva_val = base_ice = base_iva = 0.0
            impuestos = det.find('impuestos')
            if impuestos is not None:
                for imp in impuestos.findall('impuesto'):
                    try:
                        codigo = imp.find('codigo').text
                        if codigo == '3':
                            ice_val = float(imp.find('valor').text or 0)
                            base_ice = float(imp.find('baseImponible').text or 0)
                        elif codigo == '2':
                            iva_val = float(imp.find('valor').text or 0)
                            base_iva = float(imp.find('baseImponible').text or 0)
                    except Exception:
                        pass

            bot_caja = 2 if es_pack(desc) else (12 if '12U' in desc.upper() else 12)
            unidades = int(cant * bot_caja)
            info_cat = buscar_catalogo(desc)
            precio_caja = p_total / cant if cant > 0 else p_unit
            precio_bot = precio_caja / bot_caja if bot_caja > 0 else precio_caja

            reg = {
                'fecha_emision': fecha,
                'tipo_id_cliente': tipo_id, 'id_cliente': id_cli,
                'razon_social_cliente': razon,
                'codigo_producto': cod, 'nombre_producto': desc[:80],
                'codMarca': info_cat['codMarca'],
                'codProdSRI': info_cat.get('codProdSRI', ''),
                'presentacion': info_cat['presentacion'],
                'capacidad': info_cat['capacidad'], 'unidad': info_cat['unidad'],
                'grado_alcoholico': info_cat['grado'], 'codImpuesto': info_cat['codImpuesto'],
                'tipo_producto': info_cat['tipo'],
                'es_pack': es_pack(desc), 'botellas_por_caja': bot_caja,
                'cantidad_cajas': cant, 'unidades_botellas': unidades,
                'precio_unitario': p_unit, 'precio_total_sin_impuesto': p_total,
                'precio_por_caja': precio_caja, 'precio_por_botella': precio_bot,
                'base_ice': base_ice, 'valor_ice': ice_val,
                'base_iva': base_iva, 'valor_iva': iva_val,
                'importe_total': importe_total,
            }

            if ice_val > 0:
                registros_ice.append(reg.copy())
            if p_unit > 0:
                registros_pvp.append(reg.copy())

        return registros_ice, registros_pvp
    except Exception as e:
        print(f'Error parseando XML: {e}')
        return [], []


# ── Rutas ────────────────────────────────────────────────────────────────────

@ice_auditoria.route('/')
@login_required
def index():
    r = _requiere_modulo()
    if r:
        return r
    return render_template('ice_auditoria/index.html', anios=list(TAX_DB.keys()))


@ice_auditoria.route('/procesar', methods=['POST'])
@login_required
def procesar():
    r = _requiere_modulo()
    if r:
        return {'error': 'Sin acceso'}, 403

    archivos = request.files.getlist('archivos_xml')
    if not archivos:
        return {'error': 'Sube archivos XML de facturas'}, 400

    datos_ice = []
    datos_pvp = []
    archivos_ok = 0

    for arch in archivos:
        if not arch.filename.lower().endswith('.xml'):
            continue
        contenido = arch.read()
        ice, pvp = parsear_xml_ice(contenido)
        datos_ice.extend(ice)
        datos_pvp.extend(pvp)
        if ice or pvp:
            archivos_ok += 1

    return {
        'datos_ice': datos_ice,
        'datos_pvp': datos_pvp,
        'total_ice': len(datos_ice),
        'total_pvp': len(datos_pvp),
        'archivos': archivos_ok
    }


@ice_auditoria.route('/generar_xml_anexo', methods=['POST'])
@login_required
def generar_xml_anexo():
    r = _requiere_modulo()
    if r:
        return {'error': 'Sin acceso'}, 403

    data = request.get_json(force=True)
    tipo = data.get('tipo', 'ICE').upper()
    ruc = data.get('ruc', '')
    razon = data.get('razon', '')
    anio = data.get('anio', '')
    mes = data.get('mes', '')
    act_import = data.get('actImport', '02')[:2]
    datos = data.get('datos', [])

    # ── Construir entradas con codProdICE resuelto y deduplicar ──────────
    entradas_raw = []
    for reg in datos:
        info_cat = buscar_catalogo(reg.get('nombre_producto', ''))
        cod_prod_ice = _construir_cod_prod_ice(info_cat)
        tipo_id_mapped = _mapear_tipo_id(reg.get('tipo_id_cliente', '07'))

        entradas_raw.append({
            'codProdICE': cod_prod_ice,
            'gramoAzucar': '0.00',
            'tipoIdCliente': tipo_id_mapped,
            'idCliente': reg.get('id_cliente', ''),
            'tipoVentaICE': '1',
            'ventaICE': reg.get('unidades_botellas', 0),
            'devICE': '0',
            'cantProdBajaICE': '0',
        })

    # Deduplicar: combinar filas con mismo idCliente + codProdICE
    dedup = {}
    for e in entradas_raw:
        clave = (e['idCliente'], e['codProdICE'])
        if clave in dedup:
            dedup[clave]['ventaICE'] += e['ventaICE']
        else:
            dedup[clave] = dict(e)

    # ── Construir XML ────────────────────────────────────────────────────
    root_el = ET.Element(tipo.lower())
    ET.SubElement(root_el, 'TipoIDInformante').text = 'R'
    ET.SubElement(root_el, 'IdInformante').text = ruc
    ET.SubElement(root_el, 'razonSocial').text = razon
    ET.SubElement(root_el, 'Anio').text = anio
    ET.SubElement(root_el, 'Mes').text = mes
    # Orden XSD correcto: actImport ANTES de codigoOperativo
    if tipo == 'ICE':
        ET.SubElement(root_el, 'actImport').text = act_import
    ET.SubElement(root_el, 'codigoOperativo').text = tipo

    ventas = ET.SubElement(root_el, 'ventas')

    if tipo == 'ICE':
        for e in dedup.values():
            vta = ET.SubElement(ventas, 'vta')
            ET.SubElement(vta, 'codProdICE').text = e['codProdICE']
            ET.SubElement(vta, 'gramoAzucar').text = e['gramoAzucar']
            ET.SubElement(vta, 'tipoIdCliente').text = e['tipoIdCliente']
            ET.SubElement(vta, 'idCliente').text = e['idCliente']
            ET.SubElement(vta, 'tipoVentaICE').text = e['tipoVentaICE']
            ET.SubElement(vta, 'ventaICE').text = str(int(e['ventaICE']))
            ET.SubElement(vta, 'devICE').text = e['devICE']
            ET.SubElement(vta, 'cantProdBajaICE').text = e['cantProdBajaICE']
    else:
        # PVP: agrupar por id_cliente + codigo_producto
        ag = defaultdict(lambda: defaultdict(lambda: {'bot': 0, 'datos': {}}))
        for reg in datos:
            ag[reg.get('id_cliente', '')][reg.get('codigo_producto', '')]['bot'] += reg.get('unidades_botellas', 0)
            ag[reg.get('id_cliente', '')][reg.get('codigo_producto', '')]['datos'] = reg
        for id_cli, prods in sorted(ag.items()):
            for cod_p, vals in sorted(prods.items()):
                reg = vals['datos']
                vta = ET.SubElement(ventas, 'vta')
                ET.SubElement(vta, 'codProdPVP').text = reg.get('codMarca', '000001')
                ET.SubElement(vta, 'gramoAzucar').text = '0.00'
                ET.SubElement(vta, 'precioExPVP').text = f"{reg.get('precio_por_botella', 0):.2f}"
                ET.SubElement(vta, 'precioPVP').text = f"{reg.get('precio_por_botella', 0):.2f}"
                ET.SubElement(vta, 'fechaInPVP').text = f"01/{mes}/{anio}"
                ET.SubElement(vta, 'fechaFinPVP').text = f"31/{mes}/{anio}"

    xml_str = minidom.parseString(ET.tostring(root_el, encoding='utf-8')).toprettyxml(indent='  ')

    return Response(
        xml_str.encode('utf-8'),
        mimetype='application/xml',
        headers={'Content-Disposition': f'attachment; filename=Anexo_{tipo}_{ruc}_{anio}{mes}.xml'}
    )


@ice_auditoria.route('/exportar_auditoria', methods=['POST'])
@login_required
def exportar_auditoria():
    r = _requiere_modulo()
    if r:
        return {'error': 'Sin acceso'}, 403

    data = request.get_json(force=True)
    datos = data.get('datos_ice', [])
    anio = data.get('anio', '2026')
    iva_tasa_override = data.get('iva_tasa', None)

    if not datos:
        return {'error': 'No hay datos ICE para exportar'}, 400

    info_tax = TAX_DB.get(anio, TAX_DB['2026'])
    iva_tasa = iva_tasa_override if iva_tasa_override else (
        info_tax['iva'] if isinstance(info_tax['iva'], float) else 0.15
    )

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # ── Estilos ──────────────────────────────────────────────────────
        brd = Border(
            left=Side(style='thin', color='FF000000'),
            right=Side(style='thin', color='FF000000'),
            top=Side(style='thin', color='FF000000'),
            bottom=Side(style='thin', color='FF000000')
        )
        title_font = Font(name='Arial', size=14, bold=True, color='001a5276')
        h_fill = PatternFill(start_color='FF1a5276', end_color='FF1a5276', fill_type='solid')
        h_font = Font(name='Arial', size=9, bold=True, color='FFFFFFFF')
        sub_fill = PatternFill(start_color='FF2980b9', end_color='FF2980b9', fill_type='solid')
        total_fill = PatternFill(start_color='FF27ae60', end_color='FF27ae60', fill_type='solid')
        total_font = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
        nfont = Font(name='Arial', size=9, color='FF000000')
        bfont = Font(name='Arial', size=9, bold=True, color='FF000000')
        num_fmt = '#,##0.0000'
        mon_fmt = '#,##0.00'
        pct_fmt = '0.00%'

        def write_headers(ws, row, headers):
            for j, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=j, value=h)
                c.fill = h_fill
                c.font = h_font
                c.border = brd
                c.alignment = Alignment(horizontal='center', wrap_text=True)

        def apply_borders(ws, row, col_start, col_end):
            for j in range(col_start, col_end + 1):
                ws.cell(row=row, column=j).border = brd
                ws.cell(row=row, column=j).font = nfont

        wb = Workbook()

        # ================================================================
        # HOJA 1: AUDITORÍA POR PRODUCTO (con fórmulas Excel)
        # ================================================================
        ws = wb.active
        ws.title = 'Auditoría por Producto'

        ws.cell(row=1, column=1, value=f'AUDITORÍA ICE {anio} - CÁLCULO POR PRODUCTO').font = title_font
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)

        # Parámetros tributarios (referenciados en fórmulas)
        ws.cell(row=3, column=1, value='PARÁMETROS TRIBUTARIOS').font = Font(
            name='Arial', size=11, bold=True, color='00c65911')
        params = [
            ('Tarifa ICE Específica (T.Esp)', info_tax['esp']),   # B4
            ('% Ad-Valorem', 0.75),                                # B5
            ('Umbral Ad-Valorem (Umb)', info_tax['umb']),          # B6
            (f'Tasa IVA ({int(iva_tasa*100)}%)', iva_tasa),        # B7
        ]
        for i, (label, val) in enumerate(params):
            r_fila = 4 + i
            ws.cell(row=r_fila, column=1, value=label).font = bfont
            ws.cell(row=r_fila, column=1).border = brd
            c = ws.cell(row=r_fila, column=2, value=val)
            c.font = nfont
            c.number_format = num_fmt
            c.border = brd

        # Encabezados tabla
        fila = 9
        enc = [
            '#', 'Fecha', 'Cliente', 'Producto Original', 'Es Pack',
            'Producto Individual', 'Bot/Caja', 'Cajas',
            'Precio/Caja', 'Precio/Botella', 'Volumen (cc)',
            'Grado Alc. (%)', 'Precio/Litro',
            '¿Aplica AdV?', 'ICE Específico', 'ICE Ad-Valorem',
            'Total ICE', 'Base IVA', 'IVA', 'PVP Final'
        ]
        write_headers(ws, fila, enc)

        fila += 1
        fila_ini = fila
        contador = 0

        for reg in datos:
            if reg.get('es_pack'):
                productos = descomponer_pack(reg.get('nombre_producto', ''))
                num_prods = len(productos)

                for prod_nombre, prod_cap in productos:
                    info_prod = buscar_catalogo(prod_nombre)
                    cajas = float(reg.get('cantidad_cajas', 0))
                    precio_por_bot = float(reg.get('precio_total_sin_impuesto', 0)) / (num_prods * cajas) if cajas > 0 else 0
                    grado = float(info_prod.get('grado', 15))
                    vol = float(prod_cap)

                    contador += 1
                    for j, val in enumerate([
                        contador, reg.get('fecha_emision', ''),
                        (reg.get('razon_social_cliente', '') or '')[:35],
                        (reg.get('nombre_producto', '') or '')[:40], 'SÍ (PACK)',
                        f"{prod_nombre} {prod_cap}ml", 1, cajas
                    ], 1):
                        ws.cell(row=fila, column=j, value=val).font = nfont

                    ws.cell(row=fila, column=9, value=precio_por_bot).number_format = mon_fmt
                    ws.cell(row=fila, column=10, value=precio_por_bot).number_format = mon_fmt
                    ws.cell(row=fila, column=11, value=vol).font = nfont
                    ws.cell(row=fila, column=12, value=grado).font = nfont

                    # Fórmulas Excel (referencias a parámetros en B4..B7)
                    ws.cell(row=fila, column=13, value=f'=ROUND(J{fila}*1000/K{fila},4)').number_format = num_fmt
                    ws.cell(row=fila, column=14, value=f'=IF(M{fila}>$B$6,"SÍ","NO")')
                    ws.cell(row=fila, column=15, value=f'=ROUND($B$4*(L{fila}/100)*(K{fila}/1000)*G{fila}*H{fila},4)').number_format = num_fmt
                    ws.cell(row=fila, column=16, value=f'=ROUND(IF(M{fila}>$B$6,(M{fila}-$B$6)*$B$5*(K{fila}/1000)*G{fila}*H{fila},0),4)').number_format = num_fmt
                    ws.cell(row=fila, column=17, value=f'=O{fila}+P{fila}').number_format = num_fmt
                    ws.cell(row=fila, column=18, value=f'=J{fila}*G{fila}*H{fila}+Q{fila}').number_format = mon_fmt
                    ws.cell(row=fila, column=19, value=f'=ROUND(R{fila}*$B$7,4)').number_format = mon_fmt
                    ws.cell(row=fila, column=20, value=f'=R{fila}+S{fila}').number_format = mon_fmt

                    apply_borders(ws, fila, 1, 20)
                    fila += 1
            else:
                contador += 1
                vol = float(reg.get('capacidad', 750))
                grado = float(reg.get('grado_alcoholico', 15))
                cajas = float(reg.get('cantidad_cajas', 0))

                for j, val in enumerate([
                    contador, reg.get('fecha_emision', ''),
                    (reg.get('razon_social_cliente', '') or '')[:35],
                    (reg.get('nombre_producto', '') or '')[:40], 'NO',
                    (reg.get('nombre_producto', '') or '')[:40],
                    reg.get('botellas_por_caja', 12), cajas
                ], 1):
                    ws.cell(row=fila, column=j, value=val).font = nfont

                ws.cell(row=fila, column=9, value=float(reg.get('precio_por_caja', 0))).number_format = mon_fmt
                ws.cell(row=fila, column=10, value=f'=ROUND(I{fila}/G{fila},4)').number_format = mon_fmt
                ws.cell(row=fila, column=11, value=vol).font = nfont
                ws.cell(row=fila, column=12, value=grado).font = nfont

                ws.cell(row=fila, column=13, value=f'=ROUND(J{fila}*1000/K{fila},4)').number_format = num_fmt
                ws.cell(row=fila, column=14, value=f'=IF(M{fila}>$B$6,"SÍ","NO")')
                ws.cell(row=fila, column=15, value=f'=ROUND($B$4*(L{fila}/100)*(K{fila}/1000)*G{fila}*H{fila},4)').number_format = num_fmt
                ws.cell(row=fila, column=16, value=f'=ROUND(IF(M{fila}>$B$6,(M{fila}-$B$6)*$B$5*(K{fila}/1000)*G{fila}*H{fila},0),4)').number_format = num_fmt
                ws.cell(row=fila, column=17, value=f'=O{fila}+P{fila}').number_format = num_fmt
                ws.cell(row=fila, column=18, value=f'=J{fila}*G{fila}*H{fila}+Q{fila}').number_format = mon_fmt
                ws.cell(row=fila, column=19, value=f'=ROUND(R{fila}*$B$7,4)').number_format = mon_fmt
                ws.cell(row=fila, column=20, value=f'=R{fila}+S{fila}').number_format = mon_fmt

                apply_borders(ws, fila, 1, 20)
                fila += 1

        fila_fin = fila - 1

        # Fila totales
        for j in range(1, 21):
            cell = ws.cell(row=fila, column=j)
            cell.fill = total_fill
            cell.font = total_font
            cell.border = brd
        ws.cell(row=fila, column=1, value='TOTALES')
        for col in [12, 15, 16, 17, 18, 19, 20]:
            letra = get_column_letter(col)
            ws.cell(row=fila, column=col, value=f'=SUM({letra}{fila_ini}:{letra}{fila_fin})')
            ws.cell(row=fila, column=col).number_format = mon_fmt

        # Anchos
        for j, w in enumerate([5, 12, 22, 30, 10, 30, 8, 8, 12, 12, 10, 10, 12, 12, 14, 14, 12, 12, 10, 12], 1):
            ws.column_dimensions[get_column_letter(j)].width = w

        # ================================================================
        # HOJA 2: RESUMEN ICE POR PRODUCTO
        # ================================================================
        ws2 = wb.create_sheet('Resumen ICE')
        ws2.cell(row=1, column=1, value=f'RESUMEN ICE {anio}').font = title_font
        ws2.merge_cells('A1:I1')

        fila_r = 5
        enc_r = ['Producto', 'Botellas', 'Precio Total', 'ICE Específico',
                 'ICE Ad-Valorem', 'Total ICE', 'Aplica AdV?', '% ICE Esp.', '% ICE AdV.']
        write_headers(ws2, fila_r, enc_r)

        fila_r += 1
        fila_ini_r = fila_r

        prod_ag = defaultdict(lambda: {'bot': 0, 'precio': 0.0, 'ice_esp': 0.0, 'ice_adv': 0.0, 'aplica': False})
        for reg in datos:
            vol = float(reg.get('capacidad', 750))
            grado = float(reg.get('grado_alcoholico', 15))
            cajas = float(reg.get('cantidad_cajas', 0))
            precio_bot = float(reg.get('precio_por_botella', 0))

            if reg.get('es_pack'):
                prods = descomponer_pack(reg.get('nombre_producto', ''))
                num_p = len(prods)
                for pnom, pcap in prods:
                    cat = buscar_catalogo(pnom)
                    pbot = float(reg.get('precio_total_sin_impuesto', 0)) / (num_p * cajas) if cajas > 0 else 0
                    grado_p = float(cat.get('grado', 15))
                    vol_p = float(pcap)
                    i_esp = ice_especifico(info_tax['esp'], grado_p, vol_p) * cajas
                    i_adv = ice_advalorem(pbot, vol_p, info_tax['umb']) * cajas
                    k = f"{pnom} {pcap}ml (PACK)"
                    prod_ag[k]['bot'] += int(cajas)
                    prod_ag[k]['precio'] += pbot * cajas
                    prod_ag[k]['ice_esp'] += i_esp
                    prod_ag[k]['ice_adv'] += i_adv
                    p_litro = (pbot * 1000.0) / vol_p if vol_p > 0 else 0
                    prod_ag[k]['aplica'] = p_litro > info_tax['umb']
            else:
                i_esp = ice_especifico(info_tax['esp'], grado, vol) * cajas
                i_adv = ice_advalorem(precio_bot, vol, info_tax['umb']) * cajas
                k = (reg.get('nombre_producto', '') or '')[:50]
                p_litro = (precio_bot * 1000.0) / vol if vol > 0 else 0
                prod_ag[k]['bot'] += int(reg.get('unidades_botellas', 0))
                prod_ag[k]['precio'] += float(reg.get('precio_total_sin_impuesto', 0))
                prod_ag[k]['ice_esp'] += i_esp
                prod_ag[k]['ice_adv'] += i_adv
                prod_ag[k]['aplica'] = p_litro > info_tax['umb']

        for nombre, vals in sorted(prod_ag.items()):
            ice_total = vals['ice_esp'] + vals['ice_adv']
            pct_esp = vals['ice_esp'] / ice_total if ice_total > 0 else 0
            pct_adv = vals['ice_adv'] / ice_total if ice_total > 0 else 0

            for j, val in enumerate([
                nombre, vals['bot'], vals['precio'], vals['ice_esp'], vals['ice_adv'],
                ice_total, 'SÍ' if vals['aplica'] else 'NO', pct_esp, pct_adv
            ], 1):
                cell = ws2.cell(row=fila_r, column=j, value=val)
                cell.font = nfont
                cell.border = brd
                if j >= 3:
                    cell.number_format = pct_fmt if j in (8, 9) else mon_fmt
            fila_r += 1

        fila_fin_r = fila_r - 1
        for j in range(1, 10):
            cell = ws2.cell(row=fila_r, column=j)
            cell.fill = total_fill
            cell.font = total_font
            cell.border = brd
        ws2.cell(row=fila_r, column=1, value='TOTAL GENERAL')
        for col in (2, 3, 4, 5, 6):
            letra = get_column_letter(col)
            ws2.cell(row=fila_r, column=col, value=f'=SUM({letra}{fila_ini_r}:{letra}{fila_fin_r})')
            ws2.cell(row=fila_r, column=col).number_format = mon_fmt

        ws2.column_dimensions['A'].width = 45
        for j in range(2, 10):
            ws2.column_dimensions[get_column_letter(j)].width = 16

        # ================================================================
        # HOJA 3: RESUMEN GENERAL AGLUTINADO
        # ================================================================
        ws3 = wb.create_sheet('Resumen General')
        ws3.cell(row=1, column=1, value=f'RESUMEN GENERAL AGLUTINADO - ICE {anio}').font = title_font
        ws3.merge_cells('A1:H1')

        fila_g = 5
        enc_g = ['Concepto', 'Subtotal', 'ICE Específico', 'ICE Ad-Valorem',
                 'Total ICE', 'Base IVA', f'IVA {int(iva_tasa*100)}%', 'Total General']
        write_headers(ws3, fila_g, enc_g)
        fila_g += 1

        t_sub = sum(float(r.get('precio_total_sin_impuesto', 0)) for r in datos)
        t_ice_esp = sum(v['ice_esp'] for v in prod_ag.values())
        t_ice_adv = sum(v['ice_adv'] for v in prod_ag.values())
        t_ice = t_ice_esp + t_ice_adv
        b_iva = t_sub + t_ice
        iva_val = b_iva * iva_tasa
        total_gen = b_iva + iva_val

        for j, val in enumerate([
            'TOTAL VENTAS BEBIDAS ALCOHÓLICAS', t_sub, t_ice_esp,
            t_ice_adv, t_ice, b_iva, iva_val, total_gen
        ], 1):
            cell = ws3.cell(row=fila_g, column=j, value=val)
            cell.font = bfont
            cell.border = brd
            if j >= 2:
                cell.number_format = mon_fmt

        ws3.column_dimensions['A'].width = 35
        for j in range(2, 9):
            ws3.column_dimensions[get_column_letter(j)].width = 18

        # ── Guardar ──────────────────────────────────────────────────────
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=Auditoria_ICE_{anio}.xlsx'}
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {'error': str(e)}, 500
