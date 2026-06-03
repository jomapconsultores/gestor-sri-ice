"""Módulo: Facturas de Ingreso - Procesa XMLs de facturas de venta"""

from flask import Blueprint, render_template, request, Response, jsonify, flash, redirect, url_for
from flask_login import login_required, current_user
from routes.payments import usuario_tiene_modulo
from models import db
from models.user import Factura
from services.xml_parser import parse_xml_factura
from services.validaciones_sri import ValidacionesSRI
import tempfile
import os
import io
from datetime import datetime
import traceback

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

facturas_ingreso = Blueprint('facturas_ingreso', __name__)


def _requiere():
    """Verifica si el usuario tiene el módulo Facturas de Ingreso."""
    if current_user.is_admin:
        return None
    if not usuario_tiene_modulo('facturas_ingreso'):
        flash('Requieres el módulo Facturas de Ingreso ($15/mes) para acceder.', 'warning')
        return redirect(url_for('payments.ver_planes'))
    return None


@facturas_ingreso.route('/')
@login_required
def index():
    """Página principal del módulo Facturas de Ingreso."""
    r = _requiere()
    if r:
        return r

    # Listar últimas facturas procesadas del usuario
    facturas = Factura.query.filter_by(
        usuario_id=current_user.id,
        tipo='ingreso'
    ).order_by(Factura.fecha_procesamiento.desc()).limit(50).all()

    resumen = db.session.query(
        db.func.count(Factura.id).label('cantidad'),
        db.func.sum(Factura.importe_total).label('total_ventas'),
        db.func.sum(Factura.valor_ice).label('total_ice'),
        db.func.sum(Factura.valor_iva).label('total_iva'),
    ).filter_by(usuario_id=current_user.id, tipo='ingreso').first()

    return render_template('facturas_ingreso/index.html',
                           facturas=facturas,
                           resumen=resumen)


@facturas_ingreso.route('/procesar', methods=['POST'])
@login_required
def procesar():
    """Procesa múltiples XMLs de facturas de ingreso."""
    r = _requiere()
    if r:
        return jsonify({'error': 'Sin acceso'}), 403

    try:
        if 'archivos' not in request.files:
            return jsonify({'error': 'No se enviaron archivos'}), 400

        archivos = request.files.getlist('archivos')
        resultados = {'exitosas': 0, 'errores': 0, 'detalles': []}
        facturas_a_guardar = []

        for archivo in archivos:
            if not archivo.filename.lower().endswith('.xml'):
                resultados['errores'] += 1
                resultados['detalles'].append({'archivo': archivo.filename, 'error': 'No es XML'})
                continue

            try:
                # Guardar temporalmente
                with tempfile.NamedTemporaryFile(mode='wb', delete=False, suffix='.xml') as tmp:
                    archivo.save(tmp.name)
                    temp_path = tmp.name

                # Parsear
                datos = parse_xml_factura(temp_path)
                try:
                    os.unlink(temp_path)
                except:
                    pass

                if not datos:
                    resultados['errores'] += 1
                    resultados['detalles'].append({'archivo': archivo.filename, 'error': 'No se pudo parsear'})
                    continue

                # VALIDACIONES CRÍTICAS
                clave_acceso = datos.get('clave_acceso', '').strip()
                if not clave_acceso or len(clave_acceso) != 49:
                    resultados['errores'] += 1
                    resultados['detalles'].append({'archivo': archivo.filename, 'error': 'Clave inválida'})
                    continue

                # Validar RUC emisor (para facturas de VENTA)
                try:
                    ValidacionesSRI.validar_ruc(datos.get('ruc', ''))
                except ValueError as e:
                    resultados['errores'] += 1
                    resultados['detalles'].append({'archivo': archivo.filename, 'error': str(e)})
                    continue

                # Verificar si ya existe (IMPORTANTE: con usuario_id filter)
                factura_existente = Factura.query.filter_by(
                    usuario_id=current_user.id,  # ✅ CRÍTICO: Solo del usuario actual
                    clave_acceso=clave_acceso
                ).first()

                if factura_existente:
                    resultados['detalles'].append({'archivo': archivo.filename, 'estado': 'Duplicada'})
                    continue

                # Validar período fiscal
                try:
                    ValidacionesSRI.validar_periodo_fiscal(datos.get('fecha_emision', ''))
                except ValueError as e:
                    resultados['errores'] += 1
                    resultados['detalles'].append({'archivo': archivo.filename, 'error': str(e)})
                    continue

                importe_total = float(datos.get('importe_total', 0) or 0)

                # Validar importe
                try:
                    ValidacionesSRI.validar_importe(importe_total, minimo=0.01)
                except ValueError as e:
                    resultados['errores'] += 1
                    resultados['detalles'].append({'archivo': archivo.filename, 'error': str(e)})
                    continue

                # ✅ AGRUPAR IVA POR TARIFA (CRÍTICO para SRI)
                productos = datos.get('productos', [])
                iva_por_tarifa = ValidacionesSRI.agrupar_iva_por_tarifa(productos)

                # Preparar para guardar (sin commit aún)
                factura = Factura(
                    usuario_id=current_user.id,
                    clave_acceso=clave_acceso,
                    ruc_emisor=datos.get('ruc', '').strip(),
                    razon_social_emisor=datos.get('razon_social_emisor', '').strip(),
                    ruc_comprador=current_user.ruc or datos.get('id_cliente', '').strip(),
                    razon_social_comprador=current_user.nombre or datos.get('razon_social_cliente', '').strip(),
                    fecha_emision=datetime.strptime(datos.get('fecha_emision', ''), '%d/%m/%Y').date()
                        if datos.get('fecha_emision') else None,
                    numero_factura=datos.get('numero_factura', '').strip(),
                    importe_total=importe_total,
                    base_ice=sum(float(p.get('base_ice', 0) or 0) for p in productos),
                    valor_ice=sum(float(p.get('ice', 0) or 0) for p in productos),
                    # ✅ IVA AGRUPADO POR TARIFA (no suma simple)
                    base_iva=sum(iva_por_tarifa[t]['base'] for t in iva_por_tarifa),
                    valor_iva=sum(iva_por_tarifa[t]['iva'] for t in iva_por_tarifa),
                    xml_original='',
                    tipo='ingreso',
                )
                # Guardar detalles de tarifa para auditoría
                factura.notas_auditoria = f"IVA: 0%={iva_por_tarifa['0']['iva']}, 5%={iva_por_tarifa['5']['iva']}, 12%={iva_por_tarifa['12']['iva']}, 15%={iva_por_tarifa['15']['iva']}"

                facturas_a_guardar.append((factura, importe_total))
                resultados['detalles'].append({
                    'archivo': archivo.filename,
                    'estado': 'Pendiente de guardar',
                    'total': importe_total
                })

            except Exception as e:
                resultados['errores'] += 1
                resultados['detalles'].append({'archivo': archivo.filename, 'error': str(e)[:100]})

        # Guardar TODAS las facturas de una vez
        if facturas_a_guardar:
            try:
                for factura, _ in facturas_a_guardar:
                    db.session.add(factura)
                db.session.commit()
                resultados['exitosas'] = len(facturas_a_guardar)
                for i, detalle in enumerate(resultados['detalles']):
                    if detalle.get('estado') == 'Pendiente de guardar':
                        detalle['estado'] = 'Procesada'
            except Exception as e:
                db.session.rollback()
                resultados['errores'] += len(facturas_a_guardar)
                resultados['exitosas'] = 0
                for detalle in resultados['detalles']:
                    if detalle.get('estado') == 'Pendiente de guardar':
                        detalle['error'] = f'Error BD: {str(e)[:50]}'
                        del detalle['estado']

        return jsonify(resultados)

    except Exception as e:
        return jsonify({'error': str(e)}), 400


@facturas_ingreso.route('/exportar_excel', methods=['POST'])
@login_required
def exportar_excel():
    """Exporta facturas a Excel (solo valores)."""
    r = _requiere()
    if r:
        return r

    if not OPENPYXL_AVAILABLE:
        flash('Excel no disponible.', 'danger')
        return redirect(url_for('facturas_ingreso.index'))

    try:
        facturas = Factura.query.filter_by(
            usuario_id=current_user.id,
            tipo='ingreso'
        ).order_by(Factura.fecha_emision.desc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = 'Ingresos'

        # Headers
        headers = ['Fecha', 'N° Factura', 'RUC Emisor', 'Razón Social', 'Total Venta', 'Base ICE', 'ICE', 'Base IVA', 'IVA']
        ws.append(headers)

        header_fill = PatternFill(start_color='0D1B2E', end_color='0D1B2E', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col in ws[1]:
            col.fill = header_fill
            col.font = header_font

        # Datos
        for f in facturas:
            ws.append([
                f.fecha_emision.strftime('%Y-%m-%d') if f.fecha_emision else '',
                f.numero_factura or '',
                f.ruc_emisor or '',
                f.razon_social_emisor or '',
                float(f.importe_total or 0),
                float(f.base_ice or 0),
                float(f.valor_ice or 0),
                float(f.base_iva or 0),
                float(f.valor_iva or 0),
            ])

        # Totales
        ws.append(['TOTALES', '', '', ''] + [
            sum(float(f.importe_total or 0) for f in facturas),
            sum(float(f.base_ice or 0) for f in facturas),
            sum(float(f.valor_ice or 0) for f in facturas),
            sum(float(f.base_iva or 0) for f in facturas),
            sum(float(f.valor_iva or 0) for f in facturas),
        ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(output.getvalue(),
                       mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                       headers={'Content-Disposition': 'attachment; filename=Facturas_Ingreso.xlsx'})

    except Exception as e:
        flash(f'Error al generar Excel: {str(e)}', 'danger')
        return redirect(url_for('facturas_ingreso.index'))
