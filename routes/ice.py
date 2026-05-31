from flask import Blueprint, render_template, request, flash, redirect, url_for, Response
from flask_login import login_required, current_user
from services.ice_calculator import IceCalculator, TAX_DB
from services.ice_session import guardar_sesion_multiple, cargar_sesion_multiple
from routes.payments import usuario_tiene_modulo
import io
import json as _json
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

ice = Blueprint('ice', __name__)

# Mapa de períodos especiales de 2024
_PERIODOS_2024 = {
    '2024_q1': {'label': '2024 (Ene–Mar, IVA 12%)', 'iva': 0.12},
    '2024_q2': {'label': '2024 (Abr–Dic, IVA 15%)', 'iva': 0.15},
}


def _requiere_ice_simple():
    if current_user.is_admin:
        return None
    if not usuario_tiene_modulo('ice_simple'):
        flash('Requieres el módulo Cálculo ICE Simple ($10/mes) para acceder.', 'warning')
        return redirect(url_for('payments.ver_planes'))
    return None


def _requiere_ice_multiple():
    if current_user.is_admin:
        return None
    if not usuario_tiene_modulo('ice_multiple'):
        flash('Requieres el módulo ICE Múltiple + Mezcla ($15/mes) para acceder.', 'warning')
        return redirect(url_for('payments.ver_planes'))
    return None


def _resolver_anios(anios_raw):
    """
    Convierte la lista de valores del formulario en pares (etiqueta, iva_tasa).
    Maneja '2024_q1', '2024_q2' y años normales.
    """
    periodos = []
    for v in anios_raw:
        if v in _PERIODOS_2024:
            p = _PERIODOS_2024[v]
            periodos.append({'anio': '2024', 'label': p['label'], 'iva': p['iva']})
        else:
            iva = TAX_DB.get(v, {}).get('iva', 0.15)
            if iva == 'MIXTO':
                iva = 0.15
            periodos.append({'anio': v, 'label': v, 'iva': iva})
    return periodos


# ── Cálculo ICE Simple ────────────────────────────────────────────────────────

@ice.route('/calculadora')
@login_required
def calculadora():
    r = _requiere_ice_simple()
    if r:
        return r
    return render_template('ice/calculadora.html', tax_db=TAX_DB)


@ice.route('/calcular', methods=['POST'])
@login_required
def calcular():
    r = _requiere_ice_simple()
    if r:
        return r
    try:
        precio   = float(request.form.get('precio', 0))
        costos   = float(request.form.get('costos', 0))
        utilidad = float(request.form.get('utilidad', 0))
        volumen  = float(request.form.get('volumen', 750))
        grado    = float(request.form.get('grado', 35))
        tipo     = request.form.get('tipo', 'Licor')
        cantidad = int(request.form.get('cantidad', 1))
        escala   = request.form.get('escala', '')
        anios_raw = request.form.getlist('anios')

        if not anios_raw:
            flash('Selecciona al menos un período.', 'warning')
            return render_template('ice/calculadora.html', tax_db=TAX_DB)

        datos = {
            'precio_fabrica': precio,
            'volumen_cc': volumen,
            'grado_alcoholico': grado,
            'tipo_producto': tipo,
            'cantidad': cantidad,
        }
        if escala:
            datos['escala'] = escala

        periodos = _resolver_anios(anios_raw)
        resultados = {}
        for p in periodos:
            res = IceCalculator.calcular_liquidacion_completa(
                datos, p['anio'], iva_tasa=p['iva'])
            resultados[p['label']] = res

        return render_template('ice/resultado.html',
                               resultados=resultados,
                               anios_raw=anios_raw,
                               datos={
                                   'precio': precio, 'costos': costos,
                                   'utilidad': utilidad, 'volumen': volumen,
                                   'grado': grado, 'tipo': tipo, 'cantidad': cantidad,
                               })

    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('ice/calculadora.html', tax_db=TAX_DB)


# ── ICE Múltiple + Mezcla ─────────────────────────────────────────────────────

@ice.route('/multiple')
@login_required
def multiple():
    r = _requiere_ice_multiple()
    if r:
        return r
    productos_guardados = cargar_sesion_multiple(current_user.id)
    return render_template('ice/multiple.html', tax_db=TAX_DB, productos_guardados=productos_guardados)


@ice.route('/calcular_multiple', methods=['POST'])
@login_required
def calcular_multiple():
    r = _requiere_ice_multiple()
    if r:
        return r
    try:
        anios_raw = request.form.getlist('anios')
        if not anios_raw:
            flash('Selecciona al menos un período.', 'warning')
            return render_template('ice/multiple.html', tax_db=TAX_DB)

        # Leer productos del formulario
        indices = sorted({k.replace('precio_', '')
                          for k in request.form if k.startswith('precio_')})
        productos = []
        for i in indices:
            precio   = float(request.form.get(f'precio_{i}', 0) or 0)
            costos   = float(request.form.get(f'costos_{i}', 0) or 0)
            utilidad = float(request.form.get(f'utilidad_{i}', 0) or 0)
            volumen  = float(request.form.get(f'volumen_{i}', 750) or 750)
            grado    = float(request.form.get(f'grado_{i}', 35) or 35)
            tipo     = request.form.get(f'tipo_{i}', 'Licor')
            cantidad = int(request.form.get(f'cantidad_{i}', 1) or 1)
            nombre   = request.form.get(f'nombre_{i}', f'Producto {i}')
            escala   = request.form.get(f'escala_{i}', '')
            if precio <= 0 or cantidad <= 0:
                continue
            d = {
                'precio_fabrica': precio,
                'costos': costos,
                'utilidad': utilidad,
                'volumen_cc': volumen,
                'grado_alcoholico': grado,
                'tipo_producto': tipo,
                'cantidad': cantidad,
            }
            if escala:
                d['escala'] = escala
            productos.append({'nombre': nombre, 'datos': d})

        if not productos:
            flash('Agrega al menos un producto con precio y cantidad válidos.', 'warning')
            return render_template('ice/multiple.html', tax_db=TAX_DB)

        periodos = _resolver_anios(anios_raw)

        # resultados[periodo_label] = lista de dicts por producto
        resultados = {}
        totales = {}

        for p in periodos:
            lbl = p['label']
            items = []
            tot = {'ice_especifico': 0.0, 'ice_advalorem': 0.0,
                   'ice_total': 0.0, 'iva_total': 0.0, 'pvp_total': 0.0}
            for prod in productos:
                res = IceCalculator.calcular_liquidacion_completa(
                    prod['datos'], p['anio'], iva_tasa=p['iva'])
                res['nombre']   = prod['nombre']
                res['costos']   = prod['datos']['costos']
                res['utilidad'] = prod['datos']['utilidad']
                res['cantidad'] = prod['datos']['cantidad']
                # Acumular totales por cantidad
                q = prod['datos']['cantidad']
                tot['ice_especifico'] += res.get('ice_especifico_unitario', 0) * q
                tot['ice_advalorem']  += res.get('ice_advalorem_unitario', 0) * q
                tot['ice_total']      += res.get('ice_total', 0)
                tot['iva_total']      += res.get('iva_total', 0)
                tot['pvp_total']      += res.get('pvp', 0)
                items.append(res)
            resultados[lbl] = items
            totales[lbl] = {k: round(v, 4) for k, v in tot.items()}

        productos_para_guardar = [
            {
                'nombre': p['nombre'],
                'tipo_producto': p['datos'].get('tipo_producto', 'Licor'),
                'volumen_cc': p['datos'].get('volumen_cc', 750),
                'grado_alcoholico': p['datos'].get('grado_alcoholico', 35),
                'precio_fabrica': p['datos'].get('precio_fabrica', 0),
                'costos': p['datos'].get('costos', 0),
                'utilidad': p['datos'].get('utilidad', 0),
                'cantidad': p['datos'].get('cantidad', 1),
                'escala': p['datos'].get('escala', ''),
            }
            for p in productos
        ]
        guardar_sesion_multiple(current_user.id, productos_para_guardar)

        return render_template('ice/resultado_multiple.html',
                               resultados=resultados,
                               totales=totales,
                               productos=productos,
                               periodos=[p['label'] for p in periodos])

    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('ice/multiple.html', tax_db=TAX_DB)


# ── Mezcla Total ──────────────────────────────────────────────────────────────

@ice.route('/mezcla')
@login_required
def mezcla():
    r = _requiere_ice_multiple()
    if r:
        return r
    return render_template('ice/mezcla.html', tax_db=TAX_DB)


@ice.route('/calcular_mezcla', methods=['POST'])
@login_required
def calcular_mezcla():
    r = _requiere_ice_multiple()
    if r:
        return r
    try:
        anios_raw = request.form.getlist('anios')
        if not anios_raw:
            flash('Selecciona al menos un período.', 'warning')
            return render_template('ice/mezcla.html', tax_db=TAX_DB)

        indices = sorted({k.replace('precio_', '')
                          for k in request.form if k.startswith('precio_')})
        productos = []
        for i in indices:
            precio   = float(request.form.get(f'precio_{i}', 0) or 0)
            volumen  = float(request.form.get(f'volumen_{i}', 750) or 750)
            grado    = float(request.form.get(f'grado_{i}', 35) or 35)
            tipo     = request.form.get(f'tipo_{i}', 'Licor')
            cantidad = int(request.form.get(f'cantidad_{i}', 1) or 1)
            nombre   = request.form.get(f'nombre_{i}', f'Producto {i}')
            escala   = request.form.get(f'escala_{i}', '')
            if precio <= 0 or cantidad <= 0:
                continue
            d = {'precio_fabrica': precio, 'volumen_cc': volumen,
                 'grado_alcoholico': grado, 'tipo_producto': tipo, 'cantidad': cantidad}
            if escala:
                d['escala'] = escala
            productos.append({'nombre': nombre, 'datos': d})

        if not productos:
            flash('Agrega al menos un producto.', 'warning')
            return render_template('ice/mezcla.html', tax_db=TAX_DB)

        periodos = _resolver_anios(anios_raw)
        resultados = {}
        for p in periodos:
            lbl = p['label']
            items = []
            gran_total = {'ice_total': 0.0, 'iva_total': 0.0, 'pvp': 0.0}
            for prod in productos:
                res = IceCalculator.calcular_liquidacion_completa(
                    prod['datos'], p['anio'], iva_tasa=p['iva'])
                res['nombre'] = prod['nombre']
                items.append(res)
                gran_total['ice_total'] += res['ice_total']
                gran_total['iva_total'] += res['iva_total']
                gran_total['pvp']       += res['pvp']
            resultados[lbl] = {'items': items, 'total': gran_total}

        return render_template('ice/resultado_mezcla.html',
                               resultados=resultados, productos=productos)

    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('ice/mezcla.html', tax_db=TAX_DB)


# ── Exportaciones Excel ───────────────────────────────────────────────────────

@ice.route('/exportar_excel_simple', methods=['POST'])
@login_required
def exportar_excel_simple():
    r = _requiere_ice_simple()
    if r:
        return r

    if not OPENPYXL_AVAILABLE:
        flash('Excel no disponible. Contacta al administrador.', 'danger')
        return redirect(url_for('ice.calculadora'))

    try:
        precio   = float(request.form.get('precio', 0))
        volumen  = float(request.form.get('volumen', 750))
        grado    = float(request.form.get('grado', 35))
        cantidad = int(request.form.get('cantidad', 1))
        anios_json = request.form.get('anios', '[]')
        anios_raw = _json.loads(anios_json)

        if not anios_raw:
            flash('Sin períodos seleccionados.', 'warning')
            return redirect(url_for('ice.calculadora'))

        datos = {
            'precio_fabrica': precio,
            'volumen_cc': volumen,
            'grado_alcoholico': grado,
            'cantidad': cantidad,
        }

        periodos = _resolver_anios(anios_raw)

        wb = Workbook()
        ws = wb.active
        ws.title = 'ICE Simple'

        header_fill = PatternFill(start_color='0D1B2E', end_color='0D1B2E', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))

        ws['A1'] = 'Cálculo ICE Simple'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].fill = header_fill
        ws['A1'].font = Font(bold=True, color='FFFFFF', size=14)
        ws.merge_cells('A1:D1')

        ws['A2'] = f'Precio Ex-Fábrica: ${precio}'
        ws['A3'] = f'Volumen: {volumen}cc'
        ws['A4'] = f'Grado Alcohólico: {grado}%'
        ws['A5'] = f'Cantidad: {cantidad} unidad(es)'

        row = 7
        for p in periodos:
            res = IceCalculator.calcular_liquidacion_completa(datos, p['anio'], iva_tasa=p['iva'])

            ws[f'A{row}'] = p['label']
            ws[f'A{row}'].font = Font(bold=True, size=11, color='FFFFFF')
            ws[f'A{row}'].fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
            ws.merge_cells(f'A{row}:B{row}')
            row += 1

            campos = [
                ('ICE Específico Unitario', res.get('ice_especifico_unitario', 0)),
                ('ICE Ad Valorem Unitario', res.get('ice_advalorem_unitario', 0)),
                ('ICE Total Unitario', res.get('ice_total_unitario', 0)),
                ('ICE Específico Total', res.get('ice_especifico_total', 0)),
                ('ICE Ad Valorem Total', res.get('ice_advalorem_total', 0)),
                ('ICE Total', res.get('ice_total', 0)),
                ('Base IVA', res.get('base_iva', 0)),
                ('IVA Total', res.get('iva_total', 0)),
                ('PVP Final', res.get('pvp', 0)),
            ]

            for label, value in campos:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = f'${value:.2f}'
                ws[f'B{row}'].number_format = '$#,##0.00'
                row += 1

            row += 1

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(output.getvalue(),
                       mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                       headers={'Content-Disposition': 'attachment; filename=ICE_Simple.xlsx'})

    except Exception as e:
        flash(f'Error al generar Excel: {str(e)}', 'danger')
        return redirect(url_for('ice.calculadora'))


# ── Comparativa y Tarifas ─────────────────────────────────────────────────────

@ice.route('/comparativa')
@login_required
def comparativa():
    return render_template('ice/comparativa.html', tax_db=TAX_DB)


@ice.route('/tarifas')
@login_required
def ver_tarifas():
    return render_template('ice/tarifas.html', tax_db=TAX_DB)
