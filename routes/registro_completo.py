from flask import Blueprint, render_template, request, jsonify, send_file, flash, redirect, url_for, Response
from flask_login import login_required, current_user
from routes.payments import usuario_tiene_modulo
from models import db
from models.user import Factura, ClasificacionGasto
from services.xml_parser import parse_xml_factura
from services.ice_calculator import IceCalculator
from services.validaciones_sri import ValidacionesSRI
from datetime import datetime
import io, os, tempfile, traceback

registro_completo_bp = Blueprint('registro_completo', __name__)


def _requiere():
    if current_user.is_admin:
        return None
    if not usuario_tiene_modulo('registro_completo'):
        flash('Requieres el módulo Registro Completo ($120/mes).', 'warning')
        return redirect(url_for('payments.ver_planes'))
    return None


@registro_completo_bp.route('/')
@login_required
def index():
    r = _requiere()
    if r:
        return r
    return render_template('registro_completo/index.html')


@registro_completo_bp.route('/procesar_ingresos', methods=['POST'])
@login_required
def procesar_ingresos():
    r = _requiere()
    if r:
        return jsonify({'error': 'Sin acceso'}), 403
    archivos = request.files.getlist('archivos_xml')
    facturas_data = []
    for archivo in archivos:
        if not archivo.filename.lower().endswith('.xml'):
            continue
        try:
            contenido = archivo.read()
            with tempfile.NamedTemporaryFile(suffix='.xml', delete=False) as tmp:
                tmp.write(contenido)
                tmp_path = tmp.name
            datos = parse_xml_factura(tmp_path)
            os.unlink(tmp_path)
            if not datos:
                continue

            # VALIDACIONES CRÍTICAS
            ruc_emisor = datos.get('ruc', '')
            try:
                ValidacionesSRI.validar_ruc(ruc_emisor)
            except ValueError:
                continue

            fecha_str = datos.get('fecha_emision', '')
            try:
                ValidacionesSRI.validar_periodo_fiscal(fecha_str)
            except ValueError:
                continue

            importe_total = float(datos.get('importe_total', 0))
            try:
                ValidacionesSRI.validar_importe(importe_total, minimo=0.01)
            except ValueError:
                continue

            # ✅ AGRUPAR IVA POR TARIFA
            productos = datos.get('productos', [])
            iva_por_tarifa = ValidacionesSRI.agrupar_iva_por_tarifa(productos)

            facturas_data.append({
                'numero': datos.get('numero_factura', ''),
                'fecha': fecha_str,
                'ruc_emisor': ruc_emisor,
                'emisor': datos.get('razon_social_emisor', '')[:40],
                'cliente': datos.get('razon_social_cliente', '')[:40],
                'subtotal': round(
                    importe_total
                    - sum(iva_por_tarifa[t]['iva'] for t in iva_por_tarifa)
                    - float(datos.get('valor_ice_total', datos.get('valor_ice', 0))),
                    2),
                'base_ice': float(datos.get('base_ice', 0)),
                'ice': float(datos.get('valor_ice', 0)),
                'base_iva': sum(iva_por_tarifa[t]['base'] for t in iva_por_tarifa),
                'iva': sum(iva_por_tarifa[t]['iva'] for t in iva_por_tarifa),
                'iva_tarifa_0': iva_por_tarifa['0']['iva'],
                'iva_tarifa_5': iva_por_tarifa['5']['iva'],
                'iva_tarifa_12': iva_por_tarifa['12']['iva'],
                'iva_tarifa_15': iva_por_tarifa['15']['iva'],
                'total': importe_total,
            })
        except Exception as e:
            print(f"Error procesando {archivo.filename}: {traceback.format_exc()}")
            continue

    totales = {
        'subtotal':  round(sum(f['subtotal']  for f in facturas_data), 2),
        'base_ice':  round(sum(f['base_ice']  for f in facturas_data), 2),
        'ice':       round(sum(f['ice']       for f in facturas_data), 2),
        'base_iva':  round(sum(f['base_iva']  for f in facturas_data), 2),
        'iva':       round(sum(f['iva']       for f in facturas_data), 2),
        'iva_tarifa_0': round(sum(f.get('iva_tarifa_0', 0) for f in facturas_data), 2),
        'iva_tarifa_5': round(sum(f.get('iva_tarifa_5', 0) for f in facturas_data), 2),
        'iva_tarifa_12': round(sum(f.get('iva_tarifa_12', 0) for f in facturas_data), 2),
        'iva_tarifa_15': round(sum(f.get('iva_tarifa_15', 0) for f in facturas_data), 2),
        'total':     round(sum(f['total']     for f in facturas_data), 2),
    }
    return jsonify({'facturas': facturas_data, 'totales': totales, 'cantidad': len(facturas_data)})


@registro_completo_bp.route('/procesar_retenciones', methods=['POST'])
@login_required
def procesar_retenciones():
    r = _requiere()
    if r:
        return jsonify({'error': 'Sin acceso'}), 403
    archivos = request.files.getlist('archivos_xml')
    retenciones_data = []
    totales = {'renta': 0.0, 'iva': 0.0, 'isd': 0.0, 'total': 0.0}
    for archivo in archivos:
        if not archivo.filename.lower().endswith('.xml'):
            continue
        try:
            import xml.etree.ElementTree as ET
            contenido = archivo.read()
            root = ET.fromstring(contenido)

            def ft(node, tag):
                if node is None:
                    return ''
                el = node.find('.//' + tag)
                return el.text.strip() if el is not None and el.text else ''

            fecha = ft(root, 'fechaEmision') or ft(root, 'periodoFiscal')
            ruc_emisor = ft(root, 'rucEmisor') or ft(root, 'ruc')
            razon = ft(root, 'razonSocial')
            for impuesto in root.iter('impuesto'):
                codigo = ft(impuesto, 'codigo')
                base = float(ft(impuesto, 'baseImponible') or 0)
                porcentaje = float(ft(impuesto, 'porcentajeRetener') or 0)
                valor = float(ft(impuesto, 'valorRetenido') or 0)
                tipo = 'Renta' if codigo == '1' else ('IVA' if codigo == '2' else 'ISD')
                retenciones_data.append({
                    'fecha': fecha,
                    'ruc_emisor': ruc_emisor,
                    'emisor': razon[:40],
                    'tipo': tipo,
                    'base': base,
                    'porcentaje': porcentaje,
                    'valor': valor,
                })
                if tipo == 'Renta':
                    totales['renta'] += valor
                elif tipo == 'IVA':
                    totales['iva'] += valor
                else:
                    totales['isd'] += valor
                totales['total'] += valor
        except Exception:
            continue
    return jsonify({
        'retenciones': retenciones_data,
        'totales': {k: round(v, 2) for k, v in totales.items()},
        'cantidad': len(retenciones_data),
    })


@registro_completo_bp.route('/procesar_gastos', methods=['POST'])
@login_required
def procesar_gastos():
    r = _requiere()
    if r:
        return jsonify({'error': 'Sin acceso'}), 403
    archivos = request.files.getlist('archivos_xml')
    YANBAL_RUC = '1791246600001'
    gastos_data = []
    for archivo in archivos:
        if not archivo.filename.lower().endswith('.xml'):
            continue
        try:
            import tempfile
            contenido = archivo.read()
            with tempfile.NamedTemporaryFile(suffix='.xml', delete=False) as tmp:
                tmp.write(contenido)
                tmp_path = tmp.name
            datos = parse_xml_factura(tmp_path)
            os.unlink(tmp_path)
            if not datos:
                continue
            ruc = datos.get('ruc', '')
            razon = datos.get('razon_social_emisor', '')
            descuento = float(datos.get('descuento_total', 0))
            es_yanbal = ruc == YANBAL_RUC or 'YANBAL' in razon.upper()
            importe = float(datos.get('importe_total', 0))
            base_iva = float(datos.get('base_iva', 0))
            iva = float(datos.get('valor_iva', 0))
            base_imponible = base_iva - descuento if es_yanbal else base_iva
            gastos_data.append({
                'numero': datos.get('numero_factura', ''),
                'fecha': datos.get('fecha_emision', ''),
                'ruc_emisor': ruc,
                'emisor': razon[:40],
                'base_iva': base_iva,
                'iva': iva,
                'total': importe,
                'descuento': descuento,
                'base_imponible': round(base_imponible, 2),
                'tiene_descuento': descuento > 0,
                'es_yanbal': es_yanbal,
                'aplica_descuento': es_yanbal and descuento > 0,
                'categoria': 'VARIOS',
                'es_personal': False,
            })
        except Exception:
            continue
    totales = {
        'base_iva':       round(sum(g['base_iva'] for g in gastos_data), 2),
        'iva':            round(sum(g['iva'] for g in gastos_data), 2),
        'total':          round(sum(g['total'] for g in gastos_data), 2),
        'descuento_total': round(sum(g['descuento'] for g in gastos_data if g['es_yanbal']), 2),
        'base_imponible': round(sum(g['base_imponible'] for g in gastos_data if not g['es_personal']), 2),
        'iva_deducible':  round(sum(g['iva'] for g in gastos_data if not g['es_personal']), 2),
    }
    return jsonify({'gastos': gastos_data, 'totales': totales, 'cantidad': len(gastos_data)})


@registro_completo_bp.route('/procesar_ice_simple', methods=['POST'])
@login_required
def procesar_ice_simple():
    r = _requiere()
    if r:
        return jsonify({'error': 'Sin acceso'}), 403
    try:
        data = request.get_json()
        tipo_producto = data.get('tipo_producto')
        anio = data.get('anio')
        precio = float(data.get('precio', 0))
        capacidad = float(data.get('capacidad', 0))
        grado = float(data.get('grado', 0))
        cantidad = int(data.get('cantidad', 1))
        nombre = data.get('nombre', 'Producto')

        calc = IceCalculator()
        resultado = calc.calcular_liquidacion_completa(
            datos={
                'tipo_producto': tipo_producto,
                'precio_fabrica_unitario': precio,
                'volumen_cc': capacidad,
                'grado_alcoholico': grado,
                'cantidad': cantidad
            },
            anio=anio,
            iva_tasa=None
        )

        if resultado:
            return jsonify({
                'exito': True,
                'nombre': nombre,
                'tipo': tipo_producto,
                'anio': anio,
                'precio': precio,
                'capacidad': capacidad,
                'grado': grado,
                'cantidad': cantidad,
                'ice_especifico_uni': round(resultado.get('ice_especifico_unitario', 0), 2),
                'ice_advalorem_uni': round(resultado.get('ice_advalorem_unitario', 0), 2),
                'ice_total_uni': round(resultado.get('ice_total_unitario', 0), 2),
                'ice_especifico_total': round(resultado.get('ice_especifico_total', 0), 2),
                'ice_advalorem_total': round(resultado.get('ice_advalorem_total', 0), 2),
                'ice_total': round(resultado.get('ice_total', 0), 2),
                'base_iva': round(resultado.get('base_iva', 0), 2),
                'iva_uni': round(resultado.get('iva_total', 0) / cantidad, 2) if cantidad > 0 else 0,
                'iva_total': round(resultado.get('iva_total', 0), 2),
                'pvp_uni': round(resultado.get('pvp', 0), 2),
                'pvp_total': round(resultado.get('pvp', 0) * cantidad, 2),
            })
        else:
            return jsonify({'error': 'No se pudo calcular ICE para el producto.'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@registro_completo_bp.route('/procesar_ice_multiple', methods=['POST'])
@login_required
def procesar_ice_multiple():
    r = _requiere()
    if r:
        return jsonify({'error': 'Sin acceso'}), 403
    try:
        data = request.get_json()
        productos = data.get('productos', [])
        anio = data.get('anio')

        calc = IceCalculator()
        resultados = []

        for prod in productos:
            try:
                resultado = calc.calcular_liquidacion_completa(
                    datos={
                        'tipo_producto': prod.get('tipo_producto'),
                        'precio_fabrica_unitario': float(prod.get('precio', 0)),
                        'volumen_cc': float(prod.get('capacidad', 0)),
                        'grado_alcoholico': float(prod.get('grado', 0)),
                        'cantidad': int(prod.get('cantidad', 1))
                    },
                    anio=anio,
                    iva_tasa=None
                )

                if resultado:
                    cantidad = int(prod.get('cantidad', 1))
                    resultados.append({
                        'nombre': prod.get('nombre', 'Producto'),
                        'tipo': prod.get('tipo_producto'),
                        'precio': float(prod.get('precio', 0)),
                        'capacidad': float(prod.get('capacidad', 0)),
                        'grado': float(prod.get('grado', 0)),
                        'cantidad': cantidad,
                        'ice_especifico_uni': round(resultado.get('ice_especifico_unitario', 0), 2),
                        'ice_advalorem_uni': round(resultado.get('ice_advalorem_unitario', 0), 2),
                        'ice_total_uni': round(resultado.get('ice_total_unitario', 0), 2),
                        'ice_especifico_total': round(resultado.get('ice_especifico_total', 0), 2),
                        'ice_advalorem_total': round(resultado.get('ice_advalorem_total', 0), 2),
                        'ice_total': round(resultado.get('ice_total', 0), 2),
                        'base_iva': round(resultado.get('base_iva', 0), 2),
                        'iva_uni': round(resultado.get('iva_total', 0) / cantidad, 2) if cantidad > 0 else 0,
                        'iva_total': round(resultado.get('iva_total', 0), 2),
                        'pvp_uni': round(resultado.get('pvp', 0), 2),
                        'pvp_total': round(resultado.get('pvp', 0) * cantidad, 2),
                    })
            except Exception:
                continue

        return jsonify({
            'exito': True,
            'productos': resultados,
            'cantidad': len(resultados)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@registro_completo_bp.route('/exportar_excel', methods=['POST'])
@login_required
def exportar_excel():
    r = _requiere()
    if r:
        return jsonify({'error': 'Sin acceso'}), 403
    data = request.get_json(force=True)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = Workbook()
        estilos = {
            'h': {'fill': PatternFill('solid', fgColor='0D1B2E'), 'font': Font(bold=True, color='FFFFFF', size=9)},
            'n': {'font': Font(size=9)},
            't': {'fill': PatternFill('solid', fgColor='27AE60'), 'font': Font(bold=True, color='FFFFFF', size=9)},
            'border': Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin')),
        }

        def apply(cell, style_key):
            s = estilos.get(style_key, {})
            if 'fill' in s:
                cell.fill = s['fill']
            if 'font' in s:
                cell.font = s['font']
            cell.border = estilos['border']

        # Hoja 1: ICE Simple
        ws0 = wb.active
        ws0.title = 'ICE Simple'
        hdrs0 = ['Nombre', 'Precio ex-f.', 'Capacidad (L)', 'Grado (%)', 'Cantidad', 'Tipo',
                 'ICE Esp/u', 'ICE AdVal/u', 'IVA/u', 'PVP/u',
                 'ICE Esp. Total', 'ICE AdVal. Total', 'PVP Total']
        for j, h in enumerate(hdrs0, 1):
            c = ws0.cell(1, j, h); apply(c, 'h')
        for i, r in enumerate(data.get('ice_simple', []), 2):
            vals = [r.get('nombre', ''), r.get('precio', 0), r.get('capacidad', 0),
                    r.get('grado', 0), r.get('cantidad', 1), r.get('tipo', ''),
                    r.get('ice_especifico_uni', 0), r.get('ice_advalorem_uni', 0),
                    r.get('iva_uni', 0), r.get('pvp_uni', 0),
                    r.get('ice_especifico_total', 0), r.get('ice_advalorem_total', 0),
                    r.get('pvp_total', 0)]
            for j, v in enumerate(vals, 1):
                c = ws0.cell(i, j, v); apply(c, 'n')
                if j > 6: c.number_format = '#,##0.00'

        # Hoja 2: ICE Múltiple
        ws_im = wb.create_sheet('ICE Multiple')
        hdrs_im = ['Nombre', 'Precio ex-f.', 'Capacidad (L)', 'Grado (%)', 'Cantidad', 'Tipo',
                   'ICE Esp/u', 'ICE AdVal/u', 'IVA/u', 'PVP/u',
                   'ICE Esp. Total', 'ICE AdVal. Total', 'PVP Total']
        for j, h in enumerate(hdrs_im, 1):
            c = ws_im.cell(1, j, h); apply(c, 'h')
        for i, r in enumerate(data.get('ice_multiple', []), 2):
            vals = [r.get('nombre', ''), r.get('precio', 0), r.get('capacidad', 0),
                    r.get('grado', 0), r.get('cantidad', 1), r.get('tipo', ''),
                    r.get('ice_especifico_uni', 0), r.get('ice_advalorem_uni', 0),
                    r.get('iva_uni', 0), r.get('pvp_uni', 0),
                    r.get('ice_especifico_total', 0), r.get('ice_advalorem_total', 0),
                    r.get('pvp_total', 0)]
            for j, v in enumerate(vals, 1):
                c = ws_im.cell(i, j, v); apply(c, 'n')
                if j > 6: c.number_format = '#,##0.00'

        # Hoja 3: Ingresos
        ws1 = wb.create_sheet('Ingresos')
        hdrs = ['N° Factura', 'Fecha', 'RUC Emisor', 'Emisor', 'Cliente',
                'Subtotal', 'Base ICE', 'ICE', 'Base IVA', 'IVA', 'Total']
        for j, h in enumerate(hdrs, 1):
            c = ws1.cell(1, j, h)
            apply(c, 'h')
        for i, f in enumerate(data.get('ingresos', []), 2):
            vals = [f.get('numero', ''), f.get('fecha', ''), f.get('ruc_emisor', ''),
                    f.get('emisor', ''), f.get('cliente', ''), f.get('subtotal', 0),
                    f.get('base_ice', 0), f.get('ice', 0), f.get('base_iva', 0),
                    f.get('iva', 0), f.get('total', 0)]
            for j, v in enumerate(vals, 1):
                c = ws1.cell(i, j, v)
                apply(c, 'n')
                if j > 5:
                    c.number_format = '#,##0.00'

        # Hoja 4: Retenciones
        ws2 = wb.create_sheet('Retenciones')
        hdrs2 = ['Fecha', 'RUC Emisor', 'Emisor', 'Tipo', 'Base', '%', 'Valor Retenido']
        for j, h in enumerate(hdrs2, 1):
            c = ws2.cell(1, j, h)
            apply(c, 'h')
        for i, r in enumerate(data.get('retenciones', []), 2):
            vals = [r.get('fecha', ''), r.get('ruc_emisor', ''), r.get('emisor', ''),
                    r.get('tipo', ''), r.get('base', 0), r.get('porcentaje', 0), r.get('valor', 0)]
            for j, v in enumerate(vals, 1):
                c = ws2.cell(i, j, v)
                apply(c, 'n')
                if j > 3:
                    c.number_format = '#,##0.00'

        # Hoja 5: Gastos
        ws3 = wb.create_sheet('Gastos')
        hdrs3 = ['N° Factura', 'Fecha', 'RUC Emisor', 'Emisor', 'Base IVA', 'IVA',
                 'Total', 'Descuento', 'Base Imponible', 'Categoría', 'Tipo']
        for j, h in enumerate(hdrs3, 1):
            c = ws3.cell(1, j, h)
            apply(c, 'h')
        for i, g in enumerate(data.get('gastos', []), 2):
            vals = [g.get('numero', ''), g.get('fecha', ''), g.get('ruc_emisor', ''),
                    g.get('emisor', ''), g.get('base_iva', 0), g.get('iva', 0),
                    g.get('total', 0), g.get('descuento', 0), g.get('base_imponible', 0),
                    g.get('categoria', 'VARIOS'),
                    'Personal' if g.get('es_personal') else 'General']
            for j, v in enumerate(vals, 1):
                c = ws3.cell(i, j, v)
                apply(c, 'n')
                if 4 < j < 10:
                    c.number_format = '#,##0.00'

        # Hoja 6: Liquidación
        ws4 = wb.create_sheet('Liquidacion')
        liq = data.get('liquidacion', {})
        rows = [
            ('IVA en Ventas (Ingresos)', liq.get('iva_ventas', 0)),
            ('IVA en Compras (Gastos Generales)', liq.get('iva_compras', 0)),
            ('Retenciones IVA Recibidas', liq.get('ret_iva', 0)),
            ('IVA Neto', liq.get('iva_neto', 0)),
            ('', ''),
            ('ICE Específico (a declarar)', liq.get('ice_especifico', 0)),
            ('ICE Ad Valorem (75%)', liq.get('ice_advalorem', 0)),
            ('Total ICE', liq.get('total_ice', 0)),
            ('', ''),
            ('Estado IVA', liq.get('estado', '')),
            ('Crédito tributario mes anterior', liq.get('credito_anterior', 0)),
        ]
        ws4.cell(1, 1, 'LIQUIDACIÓN IVA + ICE').font = Font(bold=True, size=12, color='0D1B2E')
        for i, (label, val) in enumerate(rows, 3):
            ws4.cell(i, 1, label).font = Font(bold=True, size=9)
            c = ws4.cell(i, 2, val)
            c.font = Font(size=9)
            if isinstance(val, float):
                c.number_format = '#,##0.00'
        ws4.column_dimensions['A'].width = 35
        ws4.column_dimensions['B'].width = 18

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=Registro_Completo.xlsx'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
